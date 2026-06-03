"""
Geister-Artikel-Export: Lager-Nrn die im Portal (Stock-Analysis) gelistet sind,
aber NICHT im AMM-BESTAND existieren → akutes Verkaufs-/Storno-Risiko.

Output: Excel mit 2 Sheets:
  1. Geister-Liste (sortiert nach Storno-Schaden absteigend, mit Conditional Formatting)
  2. Zusammenfassung + Handlungsempfehlungen
"""
import pandas as pd, glob, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

USER = Path(r'C:\Users\D.Eskofier\OneDrive - elvinci.de GmbH')
BEST = r'C:\Users\D.Eskofier\Downloads\BESTAND134_20260601_2330 (1).CSV'
STOCK = r'C:\Users\D.Eskofier\Downloads\Stock-Analysis-2026-06-02T09_25_01.294Z.xlsx'
OUT = USER / 'Anlagen' / 'Geister-Artikel_Portal-vs-AMM_2026-06-02.xlsx'

# === LADEN ===
b = pd.read_csv(BEST, sep=';', encoding='ISO-8859-1', skiprows=1, header=None, low_memory=False, dtype=str)
cols=['Lager-Nr','Anzahl','Bez','LagerNr2','Lagerplatz','Klassif','Menge','WE','Notiz','Status','Auftrag','c11','c12','c13'][:b.shape[1]]
b.columns=cols
b['Lager-Nr']=b['Lager-Nr'].astype(str).str.strip()
b_set = set(b['Lager-Nr'])

s = pd.read_excel(STOCK)
s['lager_number']=s['lager_number'].astype(str).str.strip()

# Cross-Check: ist Geist evtl. schon verkauft (Sync-Verzug)?
old = pd.read_excel(sorted(glob.glob(str(USER/'elvinci - Portal ALL SOLD'/'All-Sold-*.xlsx')))[-1])
new = pd.read_excel(r'C:\Users\D.Eskofier\Downloads\All-Sold-2026-06-01T12_44_40.671Z.xlsx')
for df in [old,new]:
    df['lnr']=df['Lager Nr.'].astype(str).str.strip().str.replace(r'\.0\$','',regex=True)
combined = pd.concat([old[['lnr']], new[['lnr']]]).drop_duplicates(subset=['lnr'])
sold_set = set(combined['lnr'])

# === GEISTER FILTERN ===
ghosts = s[~s['lager_number'].isin(b_set)].copy()
ghosts['in_allsold'] = ghosts['lager_number'].isin(sold_set)
ghosts['upload_dt'] = pd.to_datetime(ghosts['datetime_upload'], errors='coerce')
NOW = datetime.datetime.now()
ghosts['alter_tage'] = (NOW - ghosts['upload_dt']).dt.days

# Empfohlene Aktion
def aktion(row):
    if row['in_allsold']:
        return 'PRÜFEN: schon verkauft (Sync-Verzug)'
    a = row['alter_tage']
    if pd.isna(a): return 'PRÜFEN: Upload-Datum fehlt'
    if a > 730: return 'AUS PORTAL ENTFERNEN (>2 Jahre alt)'
    if a > 365: return 'IM PORTAL SPERREN + KLÄREN (>1 Jahr)'
    if a > 180: return 'KLÄREN (>6 Mon. — vermutlich verloren)'
    return 'SOFORT KLÄREN: warum nicht im AMM?'
ghosts['empfehlung'] = ghosts.apply(aktion, axis=1)

# Numerische Spalten
for c in ['Buying_Price','Selling_Price','Online_Price']:
    ghosts[c] = pd.to_numeric(ghosts[c], errors='coerce').fillna(0)

# Sortieren nach Schaden (Selling_Price abs.)
ghosts = ghosts.sort_values('Selling_Price', ascending=False).reset_index(drop=True)

# === EXCEL ===
wb = Workbook()

# ==========================================
# SHEET 1: GEISTER-LISTE
# ==========================================
ws = wb.active
ws.title = 'Geister-Artikel'

# Header-Block
ws['A1'] = '🚨 GEISTER-ARTIKEL — Portal listet, AMM hat sie nicht'
ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='C00000')
ws.merge_cells('A1:K1')

ws['A2'] = f'Stand: BESTAND 01.06.2026 vs. Stock-Analysis 02.06.2026  ·  {len(ghosts)} betroffene Artikel'
ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='595959')
ws.merge_cells('A2:K2')

ws['A3'] = f'Σ Selling_Price (potenzieller Storno-Schaden): {ghosts["Selling_Price"].sum():,.2f} €  ·  Σ Buying_Price (Buchwert-Verlust): {ghosts["Buying_Price"].sum():,.2f} €'
ws['A3'].font = Font(name='Calibri', size=10, bold=True, color='0070C0')
ws.merge_cells('A3:K3')

# Spaltenkopf
headers = ['Lager-Nr','Marke','Modell','Produktgruppe','Grade',
           'EK (€)','VK (€)','Online (€)','Upload','Alter (T)','Empfohlene Aktion']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=i, value=h)
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = Border(bottom=Side(style='medium', color='000000'))

# Datenzeilen
for i, r in ghosts.iterrows():
    row = i + 6
    ws.cell(row=row, column=1,  value=r['lager_number'])
    ws.cell(row=row, column=2,  value=r['brand'] if pd.notna(r['brand']) else '')
    ws.cell(row=row, column=3,  value=str(r['model'])[:50] if pd.notna(r['model']) else '')
    ws.cell(row=row, column=4,  value=r['product_group'] if pd.notna(r['product_group']) else '')
    ws.cell(row=row, column=5,  value=r['final_grade'] if 'final_grade' in r and pd.notna(r['final_grade']) else '')
    ws.cell(row=row, column=6,  value=float(r['Buying_Price']))
    ws.cell(row=row, column=7,  value=float(r['Selling_Price']))
    ws.cell(row=row, column=8,  value=float(r['Online_Price']))
    ws.cell(row=row, column=9,  value=r['upload_dt'].date() if pd.notna(r['upload_dt']) else '')
    ws.cell(row=row, column=10, value=int(r['alter_tage']) if pd.notna(r['alter_tage']) else 0)
    ws.cell(row=row, column=11, value=r['empfehlung'])

# Formatierung Datenzeilen
n_rows = len(ghosts)
data_start = 6
data_end = data_start + n_rows - 1
for r in range(data_start, data_end + 1):
    for c in range(1, 12):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=(c == 11))
        cell.border = Border(bottom=Side(style='thin', color='D9D9D9'))
        if c in (6, 7, 8):
            cell.number_format = '#,##0.00 €'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c == 10:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif c == 9:
            cell.number_format = 'DD.MM.YYYY'
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Zebra-Streifen
for r in range(data_start, data_end + 1):
    if (r - data_start) % 2 == 1:
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor='F2F2F2')

# Conditional Formatting: hoher VK = rot, Alter > 365 = orange
red_fill = PatternFill('solid', fgColor='FFC7CE')
orange_fill = PatternFill('solid', fgColor='FFE699')
red_font = Font(color='9C0006', bold=True)

ws.conditional_formatting.add(
    f'G{data_start}:G{data_end}',
    CellIsRule(operator='greaterThan', formula=['500'], stopIfTrue=False,
               fill=red_fill, font=red_font))
ws.conditional_formatting.add(
    f'J{data_start}:J{data_end}',
    CellIsRule(operator='greaterThan', formula=['730'], stopIfTrue=False,
               fill=red_fill, font=red_font))
ws.conditional_formatting.add(
    f'J{data_start}:J{data_end}',
    CellIsRule(operator='between', formula=['365', '730'], stopIfTrue=False,
               fill=orange_fill))

# Spaltenbreiten
widths = [13, 12, 35, 25, 8, 10, 10, 10, 12, 9, 38]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 24
ws.row_dimensions[5].height = 22

# Filter + Freeze
ws.auto_filter.ref = f'A5:K{data_end}'
ws.freeze_panes = 'A6'

# ==========================================
# SHEET 2: ZUSAMMENFASSUNG
# ==========================================
ws2 = wb.create_sheet('Zusammenfassung')

ws2['A1'] = 'Zusammenfassung & Handlungsempfehlungen'
ws2['A1'].font = Font(size=16, bold=True)
ws2.merge_cells('A1:D1')

ws2['A3'] = 'METRIKEN'
ws2['A3'].font = Font(size=11, bold=True, color='1F4E79')
ws2['A3'].fill = PatternFill('solid', fgColor='DDEBF7')

metrics = [
    ('Geister-Artikel gesamt',           f'{len(ghosts)}'),
    ('davon > 1 Jahr alt im Portal',     f'{(ghosts["alter_tage"]>365).sum()} ({(ghosts["alter_tage"]>365).sum()/len(ghosts)*100:.0f}%)'),
    ('davon > 2 Jahre alt im Portal',    f'{(ghosts["alter_tage"]>730).sum()} ({(ghosts["alter_tage"]>730).sum()/len(ghosts)*100:.0f}%)'),
    ('davon schon verkauft (Sync-Verzug)', f'{ghosts["in_allsold"].sum()}'),
    ('Σ Selling_Price (Schaden bei Storno)', f'{ghosts["Selling_Price"].sum():,.2f} €'),
    ('Σ Buying_Price (Buchwert-Verlust)',    f'{ghosts["Buying_Price"].sum():,.2f} €'),
    ('Σ Online_Price (Listen-Wert)',         f'{ghosts["Online_Price"].sum():,.2f} €'),
]
for i, (k, v) in enumerate(metrics, 4):
    ws2.cell(row=i, column=1, value=k).font = Font(size=10)
    c = ws2.cell(row=i, column=2, value=v)
    c.font = Font(size=10, bold=True)
    c.alignment = Alignment(horizontal='right')

r = 4 + len(metrics) + 2
ws2.cell(row=r, column=1, value='TOP-RISIKO (>500 € Storno)').font = Font(size=11, bold=True, color='C00000')
ws2.cell(row=r, column=1).fill = PatternFill('solid', fgColor='FCE4E4')
r += 1
top = ghosts[ghosts['Selling_Price']>500].head(15)
for _, x in top.iterrows():
    ws2.cell(row=r, column=1, value=x['lager_number']).font = Font(name='Consolas', size=10)
    ws2.cell(row=r, column=2, value=f"{x['brand']} {str(x['model'])[:35]}").font = Font(size=10)
    c = ws2.cell(row=r, column=3, value=float(x['Selling_Price']))
    c.font = Font(size=10, bold=True); c.number_format='#,##0.00 €'; c.alignment=Alignment(horizontal='right')
    ws2.cell(row=r, column=4, value=f"{int(x['alter_tage'])} T").alignment=Alignment(horizontal='right')
    r += 1

r += 2
ws2.cell(row=r, column=1, value='TOP-PRODUKTGRUPPEN').font = Font(size=11, bold=True, color='1F4E79')
ws2.cell(row=r, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
r += 1
for pg, n in ghosts['product_group'].value_counts().head(8).items():
    schaden = ghosts[ghosts['product_group']==pg]['Selling_Price'].sum()
    ws2.cell(row=r, column=1, value=pg).font = Font(size=10)
    ws2.cell(row=r, column=2, value=int(n)).alignment=Alignment(horizontal='right')
    c = ws2.cell(row=r, column=3, value=float(schaden)); c.number_format='#,##0.00 €'; c.alignment=Alignment(horizontal='right')
    r += 1

r += 2
ws2.cell(row=r, column=1, value='HANDLUNGSEMPFEHLUNG').font = Font(size=11, bold=True, color='1F4E79')
ws2.cell(row=r, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
r += 1
empfehlungen = [
    '1. SOFORT: alle 65 Geister im Portal sperren (Aktion „Reservieren") bis Klärung.',
    f'2. {(ghosts["alter_tage"]>365).sum()} Artikel >1 Jahr alt → vermutlich physisch entsorgt/verloren — aus Portal entfernen.',
    f'3. {(ghosts["alter_tage"]<=180).sum()} Artikel <6 Monate alt → mit Lager-Team prüfen: Inventur-Fehler oder echter Verlust?',
    '4. Root Cause analysieren: warum verschwinden Lager-Nrn aus AMM ohne Verkauf? (Versandfehler? Manuelle Löschung? System-Glitch?)',
    '5. Prozess: nach jedem Stock-Analysis-Export einen Geister-Check (z. B. wöchentlich) — diese Liste neu erzeugen.',
]
for e in empfehlungen:
    c = ws2.cell(row=r, column=1, value=e)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.font = Font(size=10)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws2.row_dimensions[r].height = 30
    r += 1

# Spaltenbreiten Zusammenfassung
ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 10

wb.save(OUT)
print(f'Excel erstellt: {OUT}')
print(f'  Sheet 1: {len(ghosts)} Geister-Artikel · Σ VK = {ghosts["Selling_Price"].sum():,.2f} €')
print(f'  Sheet 2: Zusammenfassung + Empfehlungen')
