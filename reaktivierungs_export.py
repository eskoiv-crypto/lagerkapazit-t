"""
REAKTIVIERUNGS-LISTE — 4-Jahres-JTL-Analyse mit Stammdaten-Join
3 Sheets:
  1. 🚨 High-Value Churn (LTV>=50k, >180T inaktiv) — sofort anrufen
  2. 🌙 Reaktivierungs-Pool (LTV>=20k, 180-540T inaktiv) — Kampagne
  3. ⚠️ Pre-Churn Warning (LTV>=30k, 120-180T inaktiv) — präventiv
  4. 📊 Übersicht
"""
import pandas as pd, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

JTL = r'W:\DUSTIN EXPORTE 2026\JTL-Export-Aufträge-ab2022-08062026.csv'
STAMM = r'W:\DUSTIN EXPORTE 2026\JTL-Export-Kundenstammdaten-08062026.csv'
OUT = Path(r'C:\Users\D.Eskofier\OneDrive - elvinci.de GmbH\Anlagen') / 'Reaktivierungs-Liste_4Jahre_2026-06-08.xlsx'

print('Lade JTL 4-Jahre...')
df = pd.read_csv(JTL, sep=';', encoding='ISO-8859-1', dtype=str, low_memory=False)
df['dt'] = pd.to_datetime(df['Auftragsdatum'], errors='coerce', dayfirst=True)
df['brutto'] = pd.to_numeric(df['Brutto-VK'].str.replace(',','.'), errors='coerce').fillna(0)
df['netto']  = pd.to_numeric(df['Netto Gesamt'].str.replace(',','.'), errors='coerce').fillna(0)
df['Kunden-Nr'] = df['Kunden-Nr'].astype(str).str.strip()

print('Lade Stammdaten...')
sd = pd.read_csv(STAMM, sep=';', encoding='ISO-8859-1', dtype=str, low_memory=False)
sd['Kundennummer'] = sd['Kundennummer'].astype(str).str.strip()
sd_idx = sd.set_index('Kundennummer')

NOW = pd.Timestamp('2026-06-08')

# Pro-Kunde Aggregat
print('Aggregiere...')
kpa = df.groupby('Kunden-Nr').agg(
    n_aufträge=('Bestell Nr.','nunique'),
    n_positionen=('Bestell Nr.','count'),
    erster_kauf=('dt','min'),
    letzter_kauf=('dt','max'),
    σ_brutto=('brutto','sum'),
    σ_netto=('netto','sum'),
).reset_index()
kpa['tage_inaktiv'] = (NOW - kpa['letzter_kauf']).dt.days
kpa['jahre_aktiv'] = ((kpa['letzter_kauf'] - kpa['erster_kauf']).dt.days / 365.25).round(2)
kpa['ø_jahresumsatz'] = kpa['σ_brutto'] / kpa['jahre_aktiv'].clip(lower=0.25)

# Filter: 99999 (Sammelkunde) raus
kpa = kpa[kpa['Kunden-Nr'] != '99999'].copy()

# Top-Bezeichnung je Kunde (am häufigsten gekauft) für Angebots-Tipp
print('Top-Bezeichnungen je Kunde...')
top_bez_map = {}
for kn, grp in df[df['Kunden-Nr'] != '99999'].groupby('Kunden-Nr'):
    counts = grp['Bezeichnung'].value_counts().head(3)
    top_bez_map[kn] = ' · '.join([f"{n}× {b[:35]}" for b, n in counts.items()])
kpa['top_artikel'] = kpa['Kunden-Nr'].map(top_bez_map).fillna('—')

# Stammdaten-Join
def get_stamm(kn, col):
    try:
        v = sd_idx.loc[kn, col]
        if isinstance(v, pd.Series): v = v.iloc[0]
        return str(v) if pd.notna(v) else ''
    except (KeyError, AttributeError):
        return ''

kpa['firma'] = kpa['Kunden-Nr'].apply(lambda kn: get_stamm(kn, 'Firma'))
kpa['land'] = kpa['Kunden-Nr'].apply(lambda kn: get_stamm(kn, 'Land / ISO-Code (2-stellig)'))
kpa['ort'] = kpa['Kunden-Nr'].apply(lambda kn: get_stamm(kn, 'Ort'))
kpa['email'] = kpa['Kunden-Nr'].apply(lambda kn: get_stamm(kn, 'E-Mail-Adresse'))
kpa['telefon'] = kpa['Kunden-Nr'].apply(lambda kn: get_stamm(kn, 'Telefon'))
kpa['mobil'] = kpa['Kunden-Nr'].apply(lambda kn: get_stamm(kn, 'Mobil'))

# === SEGMENTE ===
high_value_churn = kpa[(kpa['tage_inaktiv'] > 180) & (kpa['σ_brutto'] >= 50000)].sort_values('σ_brutto', ascending=False)
reaktivierung   = kpa[(kpa['tage_inaktiv'].between(180, 540)) & (kpa['σ_brutto'].between(20000, 50000))].sort_values('σ_brutto', ascending=False)
pre_churn       = kpa[(kpa['tage_inaktiv'].between(120, 180)) & (kpa['σ_brutto'] >= 30000)].sort_values('tage_inaktiv', ascending=False)
aktiv_top       = kpa[kpa['tage_inaktiv'] < 90].sort_values('σ_brutto', ascending=False)

print(f'  High-Value Churn: {len(high_value_churn)} (Σ verlorener Wert: {high_value_churn["σ_brutto"].sum()/1000:,.0f} k €)')
print(f'  Reaktivierungs-Pool: {len(reaktivierung)} (Σ Pool-Wert: {reaktivierung["σ_brutto"].sum()/1000:,.0f} k €)')
print(f'  Pre-Churn-Warnung: {len(pre_churn)}')
print(f'  Aktive Top-Kunden: {len(aktiv_top)} (Σ aktiver Wert: {aktiv_top["σ_brutto"].sum()/1000:,.0f} k €)')

# === EXCEL ===
print('Baue Excel...')
wb = Workbook()
INK='1d1d1f'; SUBTLE='595959'; HEAD_BG='1F4E79'
RED='C00000'; AMBER='D97706'; GREEN='2da14d'; BLUE='0071e3'

COLS = [
    ('Kunden-Nr', 'Kunden-Nr', 'mono', 11),
    ('Firma', 'firma', 'text', 32),
    ('Land', 'land', 'text', 6),
    ('Ort', 'ort', 'text', 18),
    ('LTV (Σ Brutto)', 'σ_brutto', 'eur', 14),
    ('Ø Jahresumsatz', 'ø_jahresumsatz', 'eur', 14),
    ('Aufträge', 'n_aufträge', 'int', 9),
    ('Jahre aktiv', 'jahre_aktiv', 'num', 9),
    ('Erster Kauf', 'erster_kauf', 'date', 11),
    ('Letzter Kauf', 'letzter_kauf', 'date', 11),
    ('Tage inaktiv', 'tage_inaktiv', 'int', 11),
    ('Top-3 Artikel (Historie)', 'top_artikel', 'text', 50),
    ('E-Mail', 'email', 'text', 28),
    ('Telefon', 'telefon', 'text', 16),
    ('Mobil', 'mobil', 'text', 16),
    ('Status zum Abhaken', '_status', 'status', 18),
    ('Bearbeiter', '_user', 'text', 14),
    ('Datum erledigt', '_done', 'text', 13),
    ('Notiz', '_note', 'text', 30),
]

def write_sheet(ws, title, subtitle, data, header_color, status_options):
    ws['A1'] = title
    ws['A1'].font = Font(size=15, bold=True, color=header_color)
    ws.merge_cells('A1:S1')
    ws['A2'] = subtitle
    ws['A2'].font = Font(size=10, italic=True, color=SUBTLE)
    ws.merge_cells('A2:S2')
    HR = 4
    for i, (h, _, _, _) in enumerate(COLS, 1):
        c = ws.cell(row=HR, column=i, value=h)
        c.font = Font(size=10.5, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=header_color)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[HR].height = 32
    for i, (_, row) in enumerate(data.iterrows()):
        r = HR + 1 + i
        for j, (h, col, fmt, _) in enumerate(COLS, 1):
            if col == '_status':
                c = ws.cell(row=r, column=j, value='OFFEN')
                c.fill = PatternFill('solid', fgColor='FFF2CC')
                c.font = Font(size=9, bold=True, color='7F6000')
                c.alignment = Alignment(horizontal='center')
                continue
            if col.startswith('_'):
                c = ws.cell(row=r, column=j, value='')
                continue
            val = row.get(col, '')
            if isinstance(val, float) and pd.isna(val): val = ''
            if isinstance(val, pd.Timestamp): val = val.date() if not pd.isna(val) else ''
            c = ws.cell(row=r, column=j, value=val)
            c.font = Font(size=9.5)
            if fmt == 'eur':
                c.number_format = '#,##0 €'; c.alignment = Alignment(horizontal='right')
            elif fmt == 'date':
                c.number_format = 'DD.MM.YYYY'; c.alignment = Alignment(horizontal='center')
            elif fmt == 'int':
                c.alignment = Alignment(horizontal='right')
            elif fmt == 'num':
                c.number_format = '0.0'; c.alignment = Alignment(horizontal='right')
            elif fmt == 'mono':
                c.font = Font(name='Consolas', size=9.5)
            elif fmt == 'text':
                c.alignment = Alignment(wrap_text=True, vertical='center')
        # Zebra-Striping
        if i % 2 == 1:
            for j in range(1, len(COLS) + 1):
                cell = ws.cell(row=r, column=j)
                if cell.fill.fgColor.rgb in (None, '00000000', '00FFFFFF', None):
                    cell.fill = PatternFill('solid', fgColor='F8F9FA')
    # Spaltenbreiten
    for j, (_, _, _, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    if len(data):
        ws.auto_filter.ref = f'A{HR}:{get_column_letter(len(COLS))}{HR + len(data)}'
    ws.freeze_panes = f'D{HR + 1}'  # Erste 3 Spalten eingefroren

# Sheet 1: Übersicht
ws_o = wb.active
ws_o.title = 'Übersicht'
ws_o['A1'] = 'Reaktivierungs-Liste — 4-Jahres-JTL-Analyse'
ws_o['A1'].font = Font(size=18, bold=True, color=HEAD_BG)
ws_o.merge_cells('A1:E1')
ws_o['A2'] = f'Quellen: JTL-Aufträge ab 03.01.2022 · JTL-Stammdaten · Stand 08.06.2026'
ws_o['A2'].font = Font(size=10, italic=True, color=SUBTLE)
ws_o.merge_cells('A2:E2')

stats = [
    ('', '', '', '', ''),
    ('SEGMENT', 'KUNDEN', 'Σ LTV', 'Ø LTV/Kunde', 'EMPFEHLUNG'),
    ('🚨 High-Value Churn (>180T inaktiv, LTV≥50k)', len(high_value_churn), f"{high_value_churn['σ_brutto'].sum()/1000:,.0f} k €", f"{high_value_churn['σ_brutto'].mean()/1000:,.0f} k €", 'P0 — diese Woche anrufen'),
    ('🌙 Reaktivierungs-Pool (180-540T, LTV 20-50k)', len(reaktivierung), f"{reaktivierung['σ_brutto'].sum()/1000:,.0f} k €", f"{reaktivierung['σ_brutto'].mean()/1000:,.0f} k €", 'P1 — Kampagne KW 24-26'),
    ('⚠️ Pre-Churn-Warning (120-180T, LTV≥30k)', len(pre_churn), f"{pre_churn['σ_brutto'].sum()/1000:,.0f} k €", f"{pre_churn['σ_brutto'].mean()/1000:,.0f} k €", 'P1 — präventiv kontaktieren'),
    ('💎 Aktive Top-Kunden (<90T)', len(aktiv_top), f"{aktiv_top['σ_brutto'].sum()/1000:,.0f} k €", f"{aktiv_top['σ_brutto'].mean()/1000:,.0f} k €", 'pflegen, beobachten'),
    ('', '', '', '', ''),
    ('Kunden gesamt (4 Jahre)', len(kpa), f"{kpa['σ_brutto'].sum()/1000:,.0f} k €", '', ''),
]
for i, row in enumerate(stats, 4):
    for j, val in enumerate(row, 1):
        c = ws_o.cell(row=i, column=j, value=val)
        c.font = Font(size=10, bold=(i==5))
        c.alignment = Alignment(vertical='center', wrap_text=(j==1))
        if i == 5:
            c.fill = PatternFill('solid', fgColor=HEAD_BG)
            c.font = Font(size=10, bold=True, color='FFFFFF')
        elif i == 6:  # High-Value Churn
            c.fill = PatternFill('solid', fgColor='FCE8E8')
        elif i == 7:
            c.fill = PatternFill('solid', fgColor='FEF5E7')
        elif i == 8:
            c.fill = PatternFill('solid', fgColor='FFF9E6')
        elif i == 9:
            c.fill = PatternFill('solid', fgColor='E8F5E9')
    if i in (5,6,7,8,9):
        ws_o.row_dimensions[i].height = 30

# Empfehlungs-Block
ws_o.cell(row=14, column=1, value='💡 HANDLUNGSEMPFEHLUNGEN').font = Font(size=12, bold=True, color=HEAD_BG)
empf = [
    f'1. SOFORT — die {len(high_value_churn)} High-Value-Churns auf 3-4 Verkäufer verteilen, diese Woche anrufen. Σ historischer Wert: {high_value_churn["σ_brutto"].sum()/1000:,.0f} k €.',
    f'2. KW 24-26 — Reaktivierungs-Kampagne für die {len(reaktivierung)} Mid-Tier-Kunden. Personalisierte Konvolut-Angebote basierend auf "Top-3 Artikel"-Spalte.',
    f'3. PRÄVENTIV — die {len(pre_churn)} Pre-Churn-Kunden (120-180T inaktiv) bevor sie kippen. Kurzer Check-in-Anruf reicht oft.',
    f'4. PROZESS — Wöchentlicher Churn-Check als Routine: wer rutscht in den nächsten 14 T über die 180-T-Schwelle?',
    f'5. PFLEGE — die {len(aktiv_top)} aktiven Top-Kunden machen {aktiv_top["σ_brutto"].sum()/1000:,.0f} k € Σ LTV — Verlust einzelner = massiver Schaden. Account-Management einrichten.',
]
for i, e in enumerate(empf, 15):
    c = ws_o.cell(row=i, column=1, value=e); c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical='top')
    ws_o.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    ws_o.row_dimensions[i].height = 32

ws_o.column_dimensions['A'].width = 50
ws_o.column_dimensions['B'].width = 12
ws_o.column_dimensions['C'].width = 14
ws_o.column_dimensions['D'].width = 14
ws_o.column_dimensions['E'].width = 32

# Sheet 2: High-Value Churn
ws1 = wb.create_sheet('🚨 High-Value Churn')
write_sheet(ws1, f'🚨 HIGH-VALUE CHURN — {len(high_value_churn)} Kunden mit LTV ≥ 50.000 € und seit über 180 Tagen inaktiv',
            f'Σ verlorener Wert: {high_value_churn["σ_brutto"].sum()/1000:,.0f} k € · sortiert nach LTV absteigend · P0-Priorität',
            high_value_churn, RED, 'OFFEN,KONTAKTIERT,INTERESSE,KEIN_INTERESSE,ERLEDIGT')

# Sheet 3: Reaktivierungs-Pool
ws2 = wb.create_sheet('🌙 Reaktivierungs-Pool')
write_sheet(ws2, f'🌙 REAKTIVIERUNGS-POOL — {len(reaktivierung)} Kunden (LTV 20-50k, 6-18 Mon. inaktiv)',
            f'Σ Pool-Wert: {reaktivierung["σ_brutto"].sum()/1000:,.0f} k € · Kampagnen-Kandidaten · P1',
            reaktivierung, AMBER, 'OFFEN,KAMPAGNE,KONTAKTIERT,KONVOLUT_GESENDET,ERLEDIGT')

# Sheet 4: Pre-Churn
ws3 = wb.create_sheet('⚠️ Pre-Churn-Warning')
write_sheet(ws3, f'⚠️ PRE-CHURN-WARNUNG — {len(pre_churn)} Kunden (LTV ≥ 30k, 120-180 T inaktiv)',
            'PRÄVENTIV kontaktieren bevor 180-T-Schwelle erreicht ist · einfacher Check-in reicht meist',
            pre_churn, '7F6000', 'OFFEN,CHECK_IN,GESPRÄCH,ERLEDIGT')

# Sheet 5: Aktive Top
ws4 = wb.create_sheet('💎 Aktive Top-Kunden')
write_sheet(ws4, f'💎 AKTIVE TOP-KUNDEN — {len(aktiv_top)} Kunden (letzter Kauf < 90 T)',
            'Pflegen + beobachten. Account-Management bei den Top-20 etablieren.',
            aktiv_top.head(50), GREEN, 'AKTIV,KEY_ACCOUNT,REVIEW_FÄLLIG,ERLEDIGT')

wb.save(OUT)
print(f'\nExcel: {OUT}')
print(f'  Größe: {OUT.stat().st_size / 1024:.0f} KB')
