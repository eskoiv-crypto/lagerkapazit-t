"""
SCHLANK & PRÄZISE: Lager-Nrn die im AMM-BESTAND NICHT mehr geführt werden
aber im Portal (Stock-Analysis) FREI VERFÜGBAR sind.

Sortierung nach Verfügbarkeits-Status: A (aktiv) → M (manuell) → leer (vermutl. legacy)
"""
import pandas as pd, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

USER = Path(r'C:\Users\D.Eskofier\OneDrive - elvinci.de GmbH')
BEST = r'C:\Users\D.Eskofier\Downloads\BESTAND134_20260601_2330 (1).CSV'
STOCK = r'C:\Users\D.Eskofier\Downloads\Stock-Analysis-2026-06-02T09_25_01.294Z.xlsx'
OUT = USER / 'Anlagen' / 'Geister_Portal-frei-AMM-fehlt_2026-06-02.xlsx'

# Daten laden
b = pd.read_csv(BEST, sep=';', encoding='ISO-8859-1', skiprows=1, header=None, low_memory=False, dtype=str)
cols=['Lager-Nr','Anzahl','Bez','LagerNr2','Lagerplatz','Klassif','Menge','WE','Notiz','Status','Auftrag','c11','c12','c13'][:b.shape[1]]
b.columns=cols
b['Lager-Nr']=b['Lager-Nr'].astype(str).str.strip()
b_set = set(b['Lager-Nr'])

s = pd.read_excel(STOCK)
s['lager_number']=s['lager_number'].astype(str).str.strip()

# Filtern: nicht in AMM
geister = s[~s['lager_number'].isin(b_set)].copy()
geister['status_norm'] = geister['status'].fillna('leer')

# Sortier-Reihenfolge: A=1 (höchste Priorität), M=2, leer=3
geister['_sort'] = geister['status_norm'].map({'A':1,'M':2,'leer':3})
geister['upload_dt'] = pd.to_datetime(geister['datetime_upload'], errors='coerce')
NOW = datetime.datetime.now()
geister['alter_tage'] = (NOW - geister['upload_dt']).dt.days

for c in ['Buying_Price','Selling_Price','Online_Price']:
    geister[c] = pd.to_numeric(geister[c], errors='coerce').fillna(0)

geister = geister.sort_values(['_sort','Selling_Price'], ascending=[True,False]).reset_index(drop=True)

# Verfügbarkeits-Bezeichnung
status_label = {'A':'AKTIV (verkaufbar)','M':'MANUELL','leer':'(Status leer)'}
geister['Verfügbarkeit'] = geister['status_norm'].map(status_label)

# === EXCEL ===
wb = Workbook()
ws = wb.active
ws.title = 'Geister'

# Header
ws['A1'] = 'Geister: im Portal frei verfügbar, im AMM-BESTAND nicht (mehr) geführt'
ws['A1'].font = Font(size=14, bold=True, color='C00000')
ws.merge_cells('A1:I1')
ws['A2'] = f'BESTAND 01.06.2026 vs. Stock-Analysis 02.06.2026  ·  Sortiert: AKTIV → MANUELL → (leer)'
ws['A2'].font = Font(size=9, italic=True, color='595959')
ws.merge_cells('A2:I2')

# Zusammenfassung pro Verfügbarkeits-Status
ws['A4'] = 'Stufe'; ws['B4'] = 'Anzahl'; ws['C4'] = 'Σ VK (€)'; ws['D4'] = 'Σ EK (€)'
for c in 'ABCD':
    ws[f'{c}4'].font = Font(size=10, bold=True, color='FFFFFF')
    ws[f'{c}4'].fill = PatternFill('solid', fgColor='1F4E79')

row = 5
for st in ['A','M','leer']:
    sub = geister[geister['status_norm']==st]
    ws.cell(row=row, column=1, value=status_label[st])
    ws.cell(row=row, column=2, value=len(sub)).alignment = Alignment(horizontal='right')
    vk_cell = ws.cell(row=row, column=3, value=sub['Selling_Price'].sum())
    vk_cell.number_format = '#,##0.00 €'; vk_cell.alignment = Alignment(horizontal='right')
    ek_cell = ws.cell(row=row, column=4, value=sub['Buying_Price'].sum())
    ek_cell.number_format = '#,##0.00 €'; ek_cell.alignment = Alignment(horizontal='right')
    if st == 'A':
        for c in 'ABCD':
            ws[f'{c}{row}'].fill = PatternFill('solid', fgColor='C6EFCE')
            ws[f'{c}{row}'].font = Font(bold=True)
    row += 1

# Headers Tabelle
HEADER_ROW = 9
headers = ['Lager-Nr','Verfügbarkeit','Marke','Modell','Produktgruppe',
           'EK (€)','VK (€)','Online (€)','Upload','Tage']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=HEADER_ROW, column=i, value=h)
    c.font = Font(size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
    c.alignment = Alignment(horizontal='left', vertical='center')

# Datenzeilen
for i, r in geister.iterrows():
    row = HEADER_ROW + 1 + i
    ws.cell(row=row, column=1, value=r['lager_number']).font = Font(name='Consolas', size=10)
    ws.cell(row=row, column=2, value=r['Verfügbarkeit'])
    ws.cell(row=row, column=3, value=r['brand'] if pd.notna(r['brand']) else '')
    ws.cell(row=row, column=4, value=str(r['model'])[:50] if pd.notna(r['model']) else '')
    ws.cell(row=row, column=5, value=r['product_group'] if pd.notna(r['product_group']) else '')
    c6 = ws.cell(row=row, column=6, value=float(r['Buying_Price']))
    c6.number_format='#,##0.00 €'; c6.alignment=Alignment(horizontal='right')
    c7 = ws.cell(row=row, column=7, value=float(r['Selling_Price']))
    c7.number_format='#,##0.00 €'; c7.alignment=Alignment(horizontal='right')
    c8 = ws.cell(row=row, column=8, value=float(r['Online_Price']))
    c8.number_format='#,##0.00 €'; c8.alignment=Alignment(horizontal='right')
    c9 = ws.cell(row=row, column=9, value=r['upload_dt'].date() if pd.notna(r['upload_dt']) else '')
    c9.number_format='DD.MM.YYYY'; c9.alignment=Alignment(horizontal='center')
    c10 = ws.cell(row=row, column=10, value=int(r['alter_tage']) if pd.notna(r['alter_tage']) else 0)
    c10.alignment=Alignment(horizontal='right')
    # Farbiger Zeilen-Hintergrund je Verfügbarkeit
    if r['status_norm'] == 'A':
        fill = PatternFill('solid', fgColor='E2EFDA')  # grünlich
    elif r['status_norm'] == 'M':
        fill = PatternFill('solid', fgColor='FFF2CC')  # gelblich
    else:
        fill = PatternFill('solid', fgColor='F2F2F2')  # grau
    for c in range(1, 11):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = Border(bottom=Side(style='thin', color='D9D9D9'))

# Spaltenbreiten
for col, w in zip('ABCDEFGHIJ', [13,18,12,38,28,11,11,11,12,8]):
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 22
ws.row_dimensions[HEADER_ROW].height = 20

# Filter + Freeze
ws.auto_filter.ref = f'A{HEADER_ROW}:J{HEADER_ROW + len(geister)}'
ws.freeze_panes = f'A{HEADER_ROW + 1}'

wb.save(OUT)
print(f'Excel: {OUT}')
print(f'  AKTIV   (status=A):       {(geister["status_norm"]=="A").sum()} Lager-Nrn  ·  Σ VK {geister[geister["status_norm"]=="A"]["Selling_Price"].sum():,.2f} €')
print(f'  MANUELL (status=M):       {(geister["status_norm"]=="M").sum()} Lager-Nrn  ·  Σ VK {geister[geister["status_norm"]=="M"]["Selling_Price"].sum():,.2f} €')
print(f'  (Status leer):            {(geister["status_norm"]=="leer").sum()} Lager-Nrn  ·  Σ VK {geister[geister["status_norm"]=="leer"]["Selling_Price"].sum():,.2f} €')
print(f'  GESAMT:                   {len(geister)} Lager-Nrn  ·  Σ VK {geister["Selling_Price"].sum():,.2f} €')
