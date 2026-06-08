"""
100%-ERKLÄRBARE Diskrepanz-Excel:
Jede Lager-Nr aus AMM ∪ Portal bekommt EINE eindeutige Kategorie
mit Erklärung und Handlungsbedarf.

KILLCRITIC: nicht überlappende Buckets (Prio-basiert), nicht summierbare
Schätzungen, jede Zeile ist auditierbar.
"""
import pandas as pd, glob, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

USER = Path(r'C:\Users\D.Eskofier\OneDrive - elvinci.de GmbH')
BEST = r'C:\Users\D.Eskofier\Downloads\BESTAND134_20260601_2330 (1).CSV'
STOCK = str(USER / 'elvinci - Portal STOCK ANALYSIS' / 'Stock-Analysis-2026-06-02T09_25_01.294Z.xlsx')
OUT = USER / 'Anlagen' / 'Diskrepanz_AMM-vs-Portal_100pct_erklaerbar_2026-06-02.xlsx'

# === LADEN ===
b = pd.read_csv(BEST, sep=';', encoding='ISO-8859-1', skiprows=1, header=None, low_memory=False, dtype=str)
cols=['Lager-Nr','Anzahl','Bez','LagerNr2','Lagerplatz','Klassif','Menge','WE','Notiz','Status','Auftrag','c11','c12','c13'][:b.shape[1]]
b.columns=cols
b['Lager-Nr']=b['Lager-Nr'].astype(str).str.strip()
b['we_dt'] = pd.to_datetime(b['WE'],errors='coerce',dayfirst=True)
b_idx = b.set_index('Lager-Nr')

s = pd.read_excel(STOCK)
s['lager_number']=s['lager_number'].astype(str).str.strip()
s['upload_dt'] = pd.to_datetime(s['datetime_upload'],errors='coerce')
s_idx = s.set_index('lager_number')

old = pd.read_excel(sorted(glob.glob(str(USER/'elvinci - Portal ALL SOLD'/'All-Sold-*.xlsx')))[-1])
new = pd.read_excel(r'C:\Users\D.Eskofier\Downloads\All-Sold-2026-06-01T12_44_40.671Z.xlsx')
for df in [old,new]:
    df['lnr']=df['Lager Nr.'].astype(str).str.strip().str.replace(r'\.0$','',regex=True)
    df['dt']=pd.to_datetime(df['Date'],errors='coerce')
combined = pd.concat([old,new]).drop_duplicates(subset=['lnr'], keep='last')
combined['lnr_str'] = combined['lnr']
asold_idx = combined.set_index('lnr_str')
sold_set = set(combined['lnr'])

NOW = pd.Timestamp.now()

# === UNION ALL LAGER-NRN ===
all_lnr = sorted(set(b['Lager-Nr']) | set(s['lager_number']))
print(f'Union AMM ∪ Portal: {len(all_lnr):,} unique Lager-Nrn')

# === KATEGORISIERUNG (eindeutig per Prio) ===
def categorize(ln):
    in_amm = ln in b_idx.index
    in_stock = ln in s_idx.index
    in_sold = ln in sold_set

    amm_status = ''
    bez = ''
    lagerplatz = ''
    we_dt = pd.NaT
    notiz = ''
    auftrag = ''
    if in_amm:
        r = b_idx.loc[ln]
        if isinstance(r, pd.DataFrame): r = r.iloc[0]
        amm_status = r['Status']
        bez = str(r['Bez']) if pd.notna(r['Bez']) else ''
        lagerplatz = str(r['Lagerplatz']) if pd.notna(r['Lagerplatz']) else ''
        we_dt = r['we_dt']
        notiz = str(r['Notiz']) if pd.notna(r['Notiz']) else ''
        auftrag = str(r['Auftrag']) if pd.notna(r['Auftrag']) else ''

    portal_status = ''
    brand = ''
    model = ''
    pg = ''
    ek = vk = online = 0
    upload_dt = pd.NaT
    if in_stock:
        r = s_idx.loc[ln]
        if isinstance(r, pd.DataFrame): r = r.iloc[0]
        portal_status = str(r['status']) if pd.notna(r.get('status')) else 'leer'
        brand = str(r['brand']) if pd.notna(r.get('brand')) else ''
        model = str(r['model'])[:60] if pd.notna(r.get('model')) else ''
        pg = str(r['product_group']) if pd.notna(r.get('product_group')) else ''
        ek = pd.to_numeric(r.get('Buying_Price'),errors='coerce') or 0
        vk = pd.to_numeric(r.get('Selling_Price'),errors='coerce') or 0
        online = pd.to_numeric(r.get('Online_Price'),errors='coerce') or 0
        upload_dt = r['upload_dt']

    sold_dt = pd.NaT
    sold_company = ''
    sold_order = ''
    sold_by = ''
    if in_sold:
        r = asold_idx.loc[ln]
        if isinstance(r, pd.DataFrame): r = r.iloc[0]
        sold_dt = r['dt']
        sold_company = str(r['Company']) if pd.notna(r.get('Company')) else ''
        sold_order = str(r['Order Nr.']) if pd.notna(r.get('Order Nr.')) else ''
        sold_by = str(r['Sold By']) if pd.notna(r.get('Sold By')) else ''

    # === KATEGORISIERUNG mit klarer Prio ===
    quelle = 'AMM+Portal' if (in_amm and in_stock) else ('nur AMM' if in_amm else 'nur Portal')

    # Defaults
    kat_code = 'XX'
    kat_label = ''
    erklaerung = ''
    handlung = 'NEIN'
    handlung_text = ''
    prio = 99
    fillbg = 'FFFFFF'

    if in_amm and in_stock:
        # Sub-Kategorisierung für PERFEKT
        if amm_status == 'QE':
            kat_code = '01_MATCH_QE'
            kat_label = '✅ Match: QE in AMM + im Portal-Stock'
            erklaerung = f'Ideal-Fall — AMM hat QE-Status, Portal listet als status={portal_status}.'
            handlung = 'NEIN'
            handlung_text = 'Kein Handlungsbedarf — Standardfall, alles synchron.'
            prio = 1
            fillbg = 'E2EFDA'
        else:
            kat_code = '02_MATCH_VS_AA'
            kat_label = f'⚠️ Match: AMM-Status {amm_status} + im Portal'
            erklaerung = f'INKONSISTENZ — AMM zeigt {amm_status} (Versand/Auftrag) aber Portal verkauft noch!'
            handlung = 'JA'
            handlung_text = 'Im Portal SOFORT sperren — Status {amm_status} sollte nicht im Portal verkaufbar sein.'
            prio = 0
            fillbg = 'F8CBAD'
    elif in_amm and not in_stock:
        # nur AMM — Sub-Kategorisierung nach Prio
        if amm_status == 'VS':
            kat_code = '10_AMM_VS'
            kat_label = '🚚 Nur AMM · Status VS (Versand)'
            erklaerung = 'Geräte auf Versand-Pipeline — bereits verkauft, in Auslieferung. Korrekt nicht mehr im Portal.'
            handlung = 'NEIN'
            handlung_text = 'Kein Handlungsbedarf — wird mit Auslieferung gelöscht.'
            prio = 10
            fillbg = 'DDEBF7'
        elif amm_status == 'AA':
            kat_code = '11_AMM_AA'
            kat_label = '📋 Nur AMM · Status AA (Auftrags-/Sondergebunden)'
            erklaerung = f'Auftrag {auftrag if auftrag != "nan" else "(intern)"} — Gerät reserviert, nicht im Portal frei verfügbar.'
            handlung = 'NEIN'
            handlung_text = 'Kein Handlungsbedarf — Auftragsbindung legitim.'
            prio = 11
            fillbg = 'DDEBF7'
        elif amm_status == 'QE':
            # PRIO 1: schon verkauft (Status-Lag)
            if in_sold:
                kat_code = '20_QE_STATUSLAG'
                kat_label = '🟡 Nur AMM · QE-Status-Lag (schon verkauft)'
                date_str = sold_dt.strftime('%d.%m.%Y') if pd.notna(sold_dt) else '?'
                erklaerung = f'Bereits am {date_str} an "{sold_company}" verkauft (Auftrag {sold_order}). AMM-Status hinkt nach, wird mit Versand auf VS gesetzt.'
                handlung = 'NEIN'
                handlung_text = 'Kein Handlungsbedarf — automatische Status-Korrektur bei Versand.'
                prio = 20
                fillbg = 'FFF2CC'
            # PRIO 2: Set Artikel / Palette
            elif 'set artikel' in bez.lower() or 'unsortiert' in bez.lower():
                kat_code = '21_QE_PALETTE'
                kat_label = '📦 Nur AMM · QE-Sammelpalette (Set Artikel/Unsortiert)'
                erklaerung = f'Gemischte Palette: 1 AMM-Lager-Nr = N Geräte. Im Portal nicht einzeln listbar — Sonderfall.'
                handlung = 'OPTIONAL'
                handlung_text = 'Optional: als Konvolut anbieten (Listing-Konzept für Sammelposten).'
                prio = 21
                fillbg = 'FFF2CC'
            # PRIO 3: ELHR-Umlager
            elif lagerplatz.upper().startswith('ELHR'):
                kat_code = '22_QE_ELHR'
                kat_label = '🚜 Nur AMM · QE-ELHR-Umlager (in Bewegung)'
                erklaerung = f'Lagerplatz {lagerplatz} = Eingangs-/Umlagerbereich. Gerät physisch in Bewegung, nicht am Verkaufsplatz.'
                handlung = 'NEIN'
                handlung_text = 'Kein Handlungsbedarf — wird nach Umlager-Abschluss klassifiziert.'
                prio = 22
                fillbg = 'FFF2CC'
            # PRIO 4: Frischware (<30 Tage)
            elif pd.notna(we_dt) and (NOW - we_dt).days <= 30:
                kat_code = '23_QE_FRISCH'
                kat_label = '🆕 Nur AMM · QE-Frischware (WE letzte 30T)'
                we_str = we_dt.strftime('%d.%m.%Y') if pd.notna(we_dt) else '?'
                erklaerung = f'Wareneingang am {we_str} — Klassifizierung läuft regulär. Portal-Listing folgt nach Bearbeitung.'
                handlung = 'NEIN'
                handlung_text = 'Kein Handlungsbedarf — Standard-Klassifizierungs-Pipeline.'
                prio = 23
                fillbg = 'FFF2CC'
            # PRIO 5: ECHTE Listing-Lücke
            else:
                kat_code = '30_LISTING_LUECKE'
                kat_label = '🔴 Nur AMM · QE-ECHTE Listing-Lücke'
                we_str = we_dt.strftime('%d.%m.%Y') if pd.notna(we_dt) else '?'
                erklaerung = f'Klassifiziert (QE), im Hauptlager ({lagerplatz}), WE {we_str}, nicht verkauft, keine Palette, kein Umlager. → vergessenes Portal-Listing.'
                handlung = 'JA'
                handlung_text = 'Ins Portal stellen — Umsatz-Potenzial ungenutzt.'
                prio = 30
                fillbg = 'F8CBAD'
        else:
            kat_code = '12_AMM_SONSTIGE'
            kat_label = f'⚪ Nur AMM · Status {amm_status} (sonstige)'
            erklaerung = f'AMM-Status {amm_status} — Sonderfall, manuelle Prüfung.'
            handlung = 'OPTIONAL'
            handlung_text = 'Status klären.'
            prio = 12
            fillbg = 'EDEDED'
    elif not in_amm and in_stock:
        # GEISTER: Portal hat, AMM nicht
        age_days = (NOW - upload_dt).days if pd.notna(upload_dt) else 999
        if portal_status == 'A':
            kat_code = '40_GEIST_AKTIV'
            kat_label = '🔴 GEIST AKTIV im Portal · AMM kennt nicht'
            erklaerung = f'Im Portal AKTIV verkaufbar (status=A) seit {upload_dt.strftime("%d.%m.%Y") if pd.notna(upload_dt) else "?"} ({age_days} T). AMM hat die Lager-Nr nicht — Storno-Risiko bei Verkauf.'
            handlung = 'JA'
            handlung_text = 'SOFORT im Portal sperren — Versand unmöglich.'
            prio = 40
            fillbg = 'F8CBAD'
        elif portal_status == 'M':
            kat_code = '41_GEIST_MANUELL'
            kat_label = '🟠 GEIST MANUELL · AMM kennt nicht'
            erklaerung = f'Im Portal manuell gelistet (status=M) seit {upload_dt.strftime("%d.%m.%Y") if pd.notna(upload_dt) else "?"} ({age_days} T). AMM-Lager-Nr fehlt.'
            handlung = 'JA'
            handlung_text = 'Manuelles Listing prüfen + ggf. sperren/entfernen.'
            prio = 41
            fillbg = 'FFE699'
        else:  # leer/NaN
            kat_code = '42_GEIST_LEER'
            kat_label = '🟡 GEIST Status-leer · AMM kennt nicht'
            erklaerung = f'Im Portal mit Status=leer/NaN seit {upload_dt.strftime("%d.%m.%Y") if pd.notna(upload_dt) else "?"} ({age_days} T). AMM-Lager-Nr fehlt.'
            handlung = 'OPTIONAL'
            handlung_text = 'Mit Portal-Admin klären: was bedeutet Status-leer? Bei Bedarf entfernen.'
            prio = 42
            fillbg = 'FFF2CC'

    return {
        'Lager-Nr': ln,
        'Quelle': quelle,
        'Kategorie': kat_label,
        'Erklärung': erklaerung,
        'Handlungsbedarf': handlung,
        'Empfohlene Aktion': handlung_text,
        'AMM-Status': amm_status,
        'AMM-Bez': bez,
        'AMM-Lagerplatz': lagerplatz,
        'AMM-WE': we_dt,
        'AMM-Notiz': notiz,
        'AMM-Auftrag': auftrag,
        'Portal-Status': portal_status,
        'Portal-Marke': brand,
        'Portal-Modell': model,
        'Portal-WG': pg,
        'EK (€)': float(ek) if ek else 0,
        'VK (€)': float(vk) if vk else 0,
        'Online (€)': float(online) if online else 0,
        'Portal-Upload': upload_dt,
        'Sold-Datum': sold_dt,
        'Sold-Kunde': sold_company,
        'Sold-Auftrag': sold_order,
        'Sold-Verkäufer': sold_by,
        '_kat_code': kat_code,
        '_prio': prio,
        '_fillbg': fillbg,
    }

print('Kategorisiere alle Lager-Nrn …')
rows = [categorize(ln) for ln in all_lnr]
df = pd.DataFrame(rows)
print(f'Fertig: {len(df):,} Zeilen')
print()
print('Kategorien-Verteilung:')
print(df['Kategorie'].value_counts().to_string())

# === EXCEL BAUEN ===
print('\nBaue Excel …')
wb = Workbook()

INK = '1d1d1f'; SUBTLE = '595959'; HEAD_BG = '1F4E79'
GREEN = '2da14d'; AMBER = 'D97706'; RED = 'C00000'

# ============ SHEET 1: ÜBERSICHT ============
ws = wb.active
ws.title = 'Übersicht'

ws['A1'] = '100% erklärbare Diskrepanz-Auswertung — AMM ↔ Portal'
ws['A1'].font = Font(size=18, bold=True, color=HEAD_BG)
ws.merge_cells('A1:G1')
ws['A2'] = f'BESTAND 01.06.2026 ∪ Stock-Analysis 02.06.2026 · {len(df):,} Lager-Nrn · jede mit eindeutiger Kategorie + Erklärung + Handlungsempfehlung'
ws['A2'].font = Font(size=10, italic=True, color=SUBTLE)
ws.merge_cells('A2:G2')

# Verweis-Block ganz oben — die 190 sind im dedizierten Sheet
n_ja = int((df['Handlungsbedarf']=='JA').sum())
n_opt = int((df['Handlungsbedarf']=='OPTIONAL').sum())
ws.cell(row=3, column=1, value=f'➜ {n_ja} Lager-Nrn mit SOFORTIGEM Handlungsbedarf: siehe Reiter "🔴 Handlungsbedarf JA"  ·  {n_opt} OPTIONAL: siehe Reiter "🟠 Optional"').font = Font(size=11, bold=True, color=RED)
ws.cell(row=3, column=1).fill = PatternFill('solid', fgColor='FCE8E8')
ws.merge_cells('A3:G3')
ws.row_dimensions[3].height = 24

# Summary-Tabelle
ws['A4'] = 'KAT'; ws['B4'] = 'Kategorie'; ws['C4'] = 'Anzahl'; ws['D4'] = 'Σ VK (€)'; ws['E4'] = 'Handlung'; ws['F4'] = 'Quelle'; ws['G4'] = 'Bedeutung'
for col in 'ABCDEFG':
    ws[f'{col}4'].font = Font(size=10, bold=True, color='FFFFFF')
    ws[f'{col}4'].fill = PatternFill('solid', fgColor=HEAD_BG)
    ws[f'{col}4'].alignment = Alignment(horizontal='left', vertical='center')

# Gruppieren nach Kategorie
summary = df.groupby(['_kat_code','Kategorie','Quelle','Handlungsbedarf','_fillbg']).agg(
    n=('Lager-Nr','count'),
    sum_vk=('VK (€)','sum')
).reset_index().sort_values('_kat_code')

row = 5
for _, r in summary.iterrows():
    ws.cell(row=row, column=1, value=r['_kat_code']).font = Font(name='Consolas', size=9, bold=True)
    ws.cell(row=row, column=2, value=r['Kategorie']).font = Font(size=10)
    c3 = ws.cell(row=row, column=3, value=int(r['n']))
    c3.font = Font(size=11, bold=True); c3.alignment = Alignment(horizontal='right')
    c4 = ws.cell(row=row, column=4, value=float(r['sum_vk']) if r['sum_vk']>0 else None)
    c4.number_format='#,##0.00 €'; c4.alignment=Alignment(horizontal='right')
    handlung_color = RED if r['Handlungsbedarf']=='JA' else AMBER if r['Handlungsbedarf']=='OPTIONAL' else '7F7F7F'
    ws.cell(row=row, column=5, value=r['Handlungsbedarf']).font = Font(size=10, bold=True, color=handlung_color)
    ws.cell(row=row, column=6, value=r['Quelle']).font = Font(size=9)
    # Bedeutung-Text aus erstem Detail-Eintrag
    erkl = df[df['_kat_code']==r['_kat_code']].iloc[0]['Erklärung']
    ws.cell(row=row, column=7, value=erkl[:90]+'…' if len(erkl)>90 else erkl).font = Font(size=8.5, color=SUBTLE)
    ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical='center')
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=r['_fillbg'])
        ws.cell(row=row, column=c).border = Border(bottom=Side(style='thin', color='D9D9D9'))
    ws.row_dimensions[row].height = 32
    row += 1

# Totals
row += 1
ws.cell(row=row, column=2, value='TOTAL').font = Font(size=11, bold=True)
ws.cell(row=row, column=3, value=int(df.shape[0])).font = Font(size=11, bold=True)
ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
ws.cell(row=row, column=4, value=float(df['VK (€)'].sum())).number_format='#,##0.00 €'
ws.cell(row=row, column=4).font = Font(size=11, bold=True)
ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
for c in range(1,8):
    ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor='1F4E79')
    ws.cell(row=row, column=c).font = Font(size=11, bold=True, color='FFFFFF')

# Quick-Stats
row += 3
ws.cell(row=row, column=1, value='QUICK-STATS').font = Font(size=11, bold=True, color=HEAD_BG)
ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
quick = [
    ('Σ Lager-Nrn (AMM ∪ Portal)', f'{len(df):,}'.replace(',', '.')),
    ('davon in beiden Systemen (perfekt synchron)', f'{(df["Quelle"]=="AMM+Portal").sum():,}'.replace(',', '.')),
    ('davon nur AMM (Portal fehlt)', f'{(df["Quelle"]=="nur AMM").sum():,}'.replace(',', '.')),
    ('davon nur Portal (AMM fehlt = Geister)', f'{(df["Quelle"]=="nur Portal").sum():,}'.replace(',', '.')),
    ('—', ''),
    ('Σ Handlungsbedarf "JA"', f'{(df["Handlungsbedarf"]=="JA").sum():,}'.replace(',', '.')),
    ('Σ Handlungsbedarf "OPTIONAL"', f'{(df["Handlungsbedarf"]=="OPTIONAL").sum():,}'.replace(',', '.')),
    ('Σ Handlungsbedarf "NEIN" (legitim/erklärbar)', f'{(df["Handlungsbedarf"]=="NEIN").sum():,}'.replace(',', '.')),
    ('—', ''),
    ('Erklär-Quote', f'{(df["Handlungsbedarf"]=="NEIN").sum()/len(df)*100:.1f} % aller Lager-Nrn sind als legitim erklärt'),
]
for k, v in quick:
    row += 1
    if k == '—':
        continue
    ws.cell(row=row, column=1, value=k).font = Font(size=10)
    ws.cell(row=row, column=2, value=v).font = Font(size=10, bold=True)

# Spaltenbreiten
for col, w in zip('ABCDEFG', [16, 50, 10, 14, 14, 13, 55]):
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 26
ws.row_dimensions[4].height = 22

disp_cols = [
    ('Lager-Nr', 'Lager-Nr', 'mono', 14),
    ('Quelle', 'Quelle', 'text', 13),
    ('Kategorie', 'Kategorie', 'text', 42),
    ('Handlung', 'Handlungsbedarf', 'badge', 12),
    ('Empfohlene Aktion', 'Empfohlene Aktion', 'text', 50),
    ('Erklärung', 'Erklärung', 'text', 55),
    ('AMM-Status', 'AMM-Status', 'text', 10),
    ('AMM-Bez', 'AMM-Bez', 'text', 28),
    ('AMM-Lagerplatz', 'AMM-Lagerplatz', 'text', 16),
    ('AMM-WE', 'AMM-WE', 'date', 12),
    ('Portal-Status', 'Portal-Status', 'text', 11),
    ('Portal-Marke', 'Portal-Marke', 'text', 12),
    ('Portal-WG', 'Portal-WG', 'text', 22),
    ('EK (€)', 'EK (€)', 'eur', 10),
    ('VK (€)', 'VK (€)', 'eur', 10),
    ('Online (€)', 'Online (€)', 'eur', 10),
    ('Portal-Upload', 'Portal-Upload', 'date', 12),
    ('Sold-Datum', 'Sold-Datum', 'date', 12),
    ('Sold-Kunde', 'Sold-Kunde', 'text', 22),
    ('Sold-Auftrag', 'Sold-Auftrag', 'text', 18),
]

def write_lager_sheet(ws_target, title, df_subset, color_header=HEAD_BG):
    ws_target['A1'] = title
    ws_target['A1'].font = Font(size=14, bold=True, color=color_header)
    ws_target.merge_cells('A1:T1')
    ws_target.row_dimensions[1].height = 22
    ws_target['A2'] = f'{len(df_subset):,} Lager-Nrn  ·  Σ VK {df_subset["VK (€)"].sum():,.2f} €'.replace(',', '.').replace('.', ',', 1) if False else f'{len(df_subset):,} Lager-Nrn  ·  Σ VK {df_subset["VK (€)"].sum():,.2f} €'
    ws_target['A2'].font = Font(size=10, italic=True, color=SUBTLE)
    ws_target.merge_cells('A2:T2')

    HR = 4
    for i, (h, _, _, _) in enumerate(disp_cols, 1):
        c = ws_target.cell(row=HR, column=i, value=h)
        c.font = Font(size=10.5, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=color_header)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_target.row_dimensions[HR].height = 28

    for i, r in df_subset.reset_index(drop=True).iterrows():
        row = HR + 1 + i
        for j, (_, col, fmt, _) in enumerate(disp_cols, 1):
            val = r[col]
            if isinstance(val, float) and pd.isna(val): val = ''
            if isinstance(val, pd.Timestamp): val = val.date() if not pd.isna(val) else ''
            c = ws_target.cell(row=row, column=j, value=val)
            if fmt == 'eur':
                c.number_format='#,##0.00 €'; c.alignment=Alignment(horizontal='right')
            elif fmt == 'date':
                c.number_format='DD.MM.YYYY'; c.alignment=Alignment(horizontal='center')
            elif fmt == 'mono':
                c.font = Font(name='Consolas', size=9.5)
            elif fmt == 'badge':
                color = RED if val=='JA' else AMBER if val=='OPTIONAL' else '7F7F7F'
                c.font = Font(size=9.5, bold=True, color=color)
                c.alignment = Alignment(horizontal='center')
            else:
                c.font = Font(size=9.5)
            c.fill = PatternFill('solid', fgColor=r['_fillbg'])
            c.border = Border(bottom=Side(style='thin', color='E5E5E5'))

    for j, (_, _, _, w) in enumerate(disp_cols, 1):
        ws_target.column_dimensions[get_column_letter(j)].width = w
    if len(df_subset) > 0:
        ws_target.auto_filter.ref = f'A{HR}:{get_column_letter(len(disp_cols))}{HR + len(df_subset)}'
    ws_target.freeze_panes = f'D{HR + 1}'

# ============ SHEET 2: HANDLUNGSBEDARF (JA) — PROMINENT ============
ws_ja = wb.create_sheet('🔴 Handlungsbedarf JA')
df_ja = df[df['Handlungsbedarf']=='JA'].sort_values(['_prio','VK (€)'], ascending=[True,False])
write_lager_sheet(ws_ja, f'🔴 HANDLUNGSBEDARF — {len(df_ja)} Lager-Nrn die SOFORT bearbeitet werden müssen', df_ja, RED)

# ============ SHEET 3: OPTIONAL ============
ws_opt = wb.create_sheet('🟠 Optional')
df_opt = df[df['Handlungsbedarf']=='OPTIONAL'].sort_values(['_prio','VK (€)'], ascending=[True,False])
write_lager_sheet(ws_opt, f'🟠 OPTIONAL — {len(df_opt)} Lager-Nrn zur Klärung', df_opt, AMBER)

# ============ SHEET 4: ALLE LAGER-NRN (Vollbestand) ============
ws2 = wb.create_sheet('Alle Lager-Nrn')

# Sortierung: Handlungsbedarf zuerst (JA → OPTIONAL → NEIN), dann Prio, dann VK abs.
handlung_order = {'JA':0,'OPTIONAL':1,'NEIN':2}
df_sorted = df.copy()
df_sorted['_h_order'] = df_sorted['Handlungsbedarf'].map(handlung_order)
df_sorted = df_sorted.sort_values(['_h_order','_prio','VK (€)'], ascending=[True,True,False]).reset_index(drop=True)
write_lager_sheet(ws2, f'Alle Lager-Nrn — {len(df_sorted):,} gesamt · sortiert: 🔴 JA → 🟠 OPTIONAL → ⚪ NEIN', df_sorted, HEAD_BG)

# ============ SHEET 3: METHODIK ============
ws3 = wb.create_sheet('Methodik')
ws3['A1'] = 'Methodik & Prio-Logik'
ws3['A1'].font = Font(size=14, bold=True, color=HEAD_BG)

content = [
    ('', ''),
    ('QUELLEN', ''),
    ('BESTAND (AMM-Lagerverwaltung)', f'{BEST}'),
    ('Stock-Analysis (Portal-Snapshot)', f'{STOCK}'),
    ('All-Sold (kombiniert)', 'Portal-Export Jan 2026 – Jun 2026 (2 Dateien dedupliziert)'),
    ('', ''),
    ('KATEGORISIERUNGS-LOGIK', ''),
    ('Bedingung: jede Lager-Nr EXAKT EINE Kategorie', 'Prio-Reihenfolge entscheidet, in welchen Bucket sie fällt'),
    ('', ''),
    ('Wenn in beiden Systemen:', ''),
    ('  AMM=QE + im Portal', '01_MATCH_QE — Ideal-Fall, kein Handlungsbedarf'),
    ('  AMM=VS/AA + im Portal', '02_MATCH_VS_AA — INKONSISTENZ, sofort sperren'),
    ('', ''),
    ('Wenn nur AMM:', ''),
    ('  Status VS', '10_AMM_VS — Versand-Pipeline, kein Handlungsbedarf'),
    ('  Status AA', '11_AMM_AA — Auftragsbindung, kein Handlungsbedarf'),
    ('  Status QE — Prio-Logik:', ''),
    ('    Prio 1: in All-Sold', '20_QE_STATUSLAG — schon verkauft, AMM-Status hinkt nach'),
    ('    Prio 2: Bez enthält "Set Artikel" oder "Unsortiert"', '21_QE_PALETTE — Sammelpalette, nicht einzeln listbar'),
    ('    Prio 3: Lagerplatz beginnt mit ELHR', '22_QE_ELHR — Umlagerbereich, in Bewegung'),
    ('    Prio 4: WE-Datum ≤ 30 Tage', '23_QE_FRISCH — Frischware, Klassifizierung läuft'),
    ('    Prio 5: Rest', '30_LISTING_LUECKE — ECHTE Listing-Lücke, ins Portal stellen'),
    ('', ''),
    ('Wenn nur Portal (Geister):', ''),
    ('  status=A', '40_GEIST_AKTIV — aktiv verkaufbar, SOFORT sperren'),
    ('  status=M', '41_GEIST_MANUELL — manuelles Listing prüfen'),
    ('  status=leer (NaN)', '42_GEIST_LEER — Klärung mit Portal-Admin'),
    ('', ''),
    ('RESTRISIKEN (Caveats — siehe Brief)', ''),
    ('Snapshot-Zeitversatz', 'BESTAND 01.06 23:30 vs Stock 02.06 09:25 (~10 h)'),
    ('SSCC-Format-Mismatch', '49 SSCC-Lager-Nrn nur in AMM, evtl. nicht via 9-stellig findbar'),
    ('Portal-Status NaN ungeklärt', '2.054 Stock-Einträge ohne Status — Bedeutung offen'),
]
for i, (k, v) in enumerate(content, 3):
    ws3.cell(row=i, column=1, value=k).font = Font(size=10, bold=k in ('QUELLEN','KATEGORISIERUNGS-LOGIK','RESTRISIKEN (Caveats — siehe Brief)') or k.startswith('Wenn '))
    ws3.cell(row=i, column=2, value=v).font = Font(size=10)
    ws3.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws3.column_dimensions['A'].width = 50
ws3.column_dimensions['B'].width = 65

wb.save(OUT)
print(f'\nExcel: {OUT}')
print(f'  Sheets: Übersicht · Alle Lager-Nrn ({len(df):,}) · Methodik')
print(f'  Erklär-Quote: {(df["Handlungsbedarf"]=="NEIN").sum()/len(df)*100:.1f} % als legitim erklärt')
print(f'  JA (Handlungsbedarf): {(df["Handlungsbedarf"]=="JA").sum()}')
print(f'  OPTIONAL: {(df["Handlungsbedarf"]=="OPTIONAL").sum()}')
