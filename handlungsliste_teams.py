"""
TEAM-OPTIMIERTE Handlungsliste — getrennt nach IT und Lager.
Jedes Item: nummerierte Steps · Tool · Verifikation · Status-Spalte zum Abhaken.

KILLCRITIC:
- IT vs Lager klar getrennt (kein Rätselraten)
- Step-by-Step statt Satz-Empfehlung
- Verifikations-Anweisung pro Task
- Doppelbuchungs-Track (NEU)
- Überzählig-Track (NEU)
- Status-Spalten für Tracking (OFFEN/IN ARBEIT/ERLEDIGT)
"""
import pandas as pd, glob, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

USER = Path(r'C:\Users\D.Eskofier\OneDrive - elvinci.de GmbH')
BEST = r'C:\Users\D.Eskofier\Downloads\BESTAND134_20260601_2330 (1).CSV'
STOCK = str(USER/'elvinci - Portal STOCK ANALYSIS'/'Stock-Analysis-2026-06-02T09_25_01.294Z.xlsx')
ORDER_CSV = r'C:\Users\D.Eskofier\Downloads\387598253DUSTIN_Auftrag_AU2026031399977 1 (1).csv'
UZ = r'C:\Users\D.Eskofier\Downloads\Überzählig (1).xlsx'
OUT = USER / 'Anlagen' / 'Handlungsliste_IT-Lager_Teams_2026-06-02_v2-SSCC.xlsx'

# === LADEN ===
b = pd.read_csv(BEST, sep=';', encoding='ISO-8859-1', skiprows=1, header=None, low_memory=False, dtype=str)
cols=['Lager-Nr','Anzahl','Bez','LagerNr2','Lagerplatz','Klassif','Menge','WE','Notiz','Status','Auftrag','c11','c12','c13'][:b.shape[1]]
b.columns=cols
b['Lager-Nr']=b['Lager-Nr'].astype(str).str.strip()
b['we_dt'] = pd.to_datetime(b['WE'],errors='coerce',dayfirst=True)
b_set = set(b['Lager-Nr'])
b_lnr2_set = set(b['LagerNr2'].astype(str).str.strip()) - {'', 'nan'}
# Reverse-Lookup: LagerNr2 → (primary Lager-Nr + Status + Bez + Lagerplatz)
b_sscc_lookup = {}
for _, r in b.iterrows():
    ln2 = str(r['LagerNr2']).strip()
    if ln2 and ln2 != 'nan':
        b_sscc_lookup[ln2] = (r['Lager-Nr'], r['Status'], r['Bez'], r['Lagerplatz'])
b_idx = b.set_index('Lager-Nr')

s = pd.read_excel(STOCK)
s['lager_number']=s['lager_number'].astype(str).str.strip()
s['upload_dt']=pd.to_datetime(s['datetime_upload'],errors='coerce')
s_set = set(s['lager_number'])
s_idx = s.set_index('lager_number')

old = pd.read_excel(sorted(glob.glob(str(USER/'elvinci - Portal ALL SOLD'/'All-Sold-*.xlsx')))[-1])
new = pd.read_excel(r'C:\Users\D.Eskofier\Downloads\All-Sold-2026-06-01T12_44_40.671Z.xlsx')
for df in [old,new]:
    df['lnr']=df['Lager Nr.'].astype(str).str.strip().str.replace(r'\.0$','',regex=True)
sold_set = set(pd.concat([old['lnr'],new['lnr']]).unique())

jtl = pd.read_csv(USER/'elvinci - JTL EXPORT'/'JTL-Export-Aufträge-02062026.csv', sep=';', encoding='ISO-8859-1', dtype=str, low_memory=False)
jtl['Artikelnummer'] = jtl['Artikelnummer'].astype(str).str.strip()
jtl_lnrs = set(jtl['Artikelnummer'])

ord_df = pd.read_csv(ORDER_CSV, sep=';', encoding='utf-8', dtype=str)
ord_lnrs = set(ord_df['TA Nummer'].astype(str).str.strip())

uz_df = pd.read_excel(UZ, header=None)
uz_lnrs = set()
for col in uz_df.columns:
    for v in uz_df[col].dropna():
        try:
            lnr = str(int(float(v))).strip()
            if lnr.isdigit(): uz_lnrs.add(lnr)
        except: pass

NOW = pd.Timestamp.now()

# ============================================================
# IT-TASK-LISTE bauen
# ============================================================
it_tasks = []

# IT-TASK A: Geister AKTIV — KILLCRITIC: SSCC-Check VOR Klassifikation
geister = s[~s['lager_number'].isin(b_set)].copy()
geister['status_norm'] = geister['status'].fillna('leer')
geister['upload_dt'] = pd.to_datetime(geister['datetime_upload'],errors='coerce')
geister['alter_tage'] = (NOW - geister['upload_dt']).dt.days
# NEU: SSCC-Mapping-Check über LagerNr2
geister['sscc_match'] = geister['lager_number'].isin(b_lnr2_set)
for c in ['Buying_Price','Selling_Price','Online_Price']:
    geister[c] = pd.to_numeric(geister[c],errors='coerce').fillna(0)

# Splitte: echte Geister (nicht in LagerNr2) vs SSCC-Mapping (in LagerNr2)
echte_geister = geister[~geister['sscc_match']]
sscc_mapping = geister[geister['sscc_match']]

# --- ECHTE Geister AKTIV (5) ---
for _, r in echte_geister[echte_geister['status_norm']=='A'].sort_values('Selling_Price',ascending=False).iterrows():
    in_jtl = r['lager_number'] in jtl_lnrs
    if in_jtl:
        au_list = jtl[jtl['Artikelnummer']==r['lager_number']]['Bestell Nr.'].unique().tolist()
        it_tasks.append({
            'Prio': 'P1',
            'Typ': 'AMM-LÜCKE',
            'Team': 'IT (AMM-Admin)',
            'Lager-Nr': r['lager_number'],
            'Diagnose': f'Im Portal AKTIV + in JTL-Auftrag {au_list[0]} — aber im AMM verschwunden (auch nicht als SSCC-LagerNr2)',
            'Step 1': f'Lager fragen: ist Gerät physisch da? (Marke {r["brand"]}, {str(r["model"])[:30]})',
            'Step 2': 'Wenn JA: AMM-Eintrag neu anlegen mit Status QE + Lagerplatz',
            'Step 3': f'Wenn NEIN: Portal-Sperre + JTL-Auftrag {au_list[0]} stornieren/umbuchen',
            'Tool': 'AMM-Admin-UI · JTL · Portal-Admin',
            'Verifikation': 'Nach Anlage: nächster BESTAND-Export enthält Lager-Nr · JTL-Auftrag versendbar',
            'VK (€)': float(r['Selling_Price']),
            'Alter (T)': int(r['alter_tage']) if pd.notna(r['alter_tage']) else 0,
            'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
        })
    else:
        it_tasks.append({
            'Prio': 'P1',
            'Typ': 'GEIST-AKTIV',
            'Team': 'IT (Portal-Admin)',
            'Lager-Nr': r['lager_number'],
            'Diagnose': f'Im Portal AKTIV verkaufbar (status=A) seit {int(r["alter_tage"]) if pd.notna(r["alter_tage"]) else "?"} T — auch nicht als SSCC-LagerNr2 → echter Geist, Storno-Risiko',
            'Step 1': f'Lager fragen: ist {r["lager_number"]} physisch im Lager? (Bestätigung schriftlich)',
            'Step 2': 'Wenn NEIN: Portal-Listing auf BLOCKED setzen (Notiz: "Geister-Bereinigung TT.MM.JJJJ")',
            'Step 3': 'Wenn JA: AMM-Admin informieren — Lager-Nr-Anlage nötig',
            'Tool': 'Portal-Admin-Backend → Listings → Suche Lager-Nr',
            'Verifikation': 'Nach 24h: nächster Stock-Analysis-Export enthält Lager-Nr NICHT mehr',
            'VK (€)': float(r['Selling_Price']),
            'Alter (T)': int(r['alter_tage']) if pd.notna(r['alter_tage']) else 0,
            'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
        })

# --- ECHTE Geister MANUELL (3) ---
for _, r in echte_geister[echte_geister['status_norm']=='M'].sort_values('Selling_Price',ascending=False).iterrows():
    it_tasks.append({
        'Prio': 'P1',
        'Typ': 'GEIST-MANUELL',
        'Team': 'IT (Portal-Admin)',
        'Lager-Nr': r['lager_number'],
        'Diagnose': f'Manuell gelistetes Portal-Item (status=M) seit {int(r["alter_tage"]) if pd.notna(r["alter_tage"]) else "?"} T — kein AMM-Pendant (auch nicht als SSCC-LagerNr2)',
        'Step 1': 'Portal-Listing-Historie prüfen: wer hat manuelles Listing angelegt + warum?',
        'Step 2': 'Lager-Bestätigung physisch vorhanden?',
        'Step 3': 'Bei NEIN: Listing entfernen. Bei JA: AMM-Anlage + Status=A statt M',
        'Tool': 'Portal-Admin · Audit-Log · AMM-Admin',
        'Verifikation': 'Listing-Historie dokumentiert + nächste Stock-Analysis konsistent',
        'VK (€)': float(r['Selling_Price']),
        'Alter (T)': int(r['alter_tage']) if pd.notna(r['alter_tage']) else 0,
        'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
    })

# --- NEU: SSCC-LISTING-VERALTET (29) ---
# Aktive + Manuell-Geister mit SSCC-Mapping = Portal listet veraltete Einzel-Lager-Nr
sscc_aktiv = sscc_mapping[sscc_mapping['status_norm'].isin(['A','M'])]
if len(sscc_aktiv) > 0:
    it_tasks.append({
        'Prio': 'P1',
        'Typ': 'SSCC-LISTING-VERALTET',
        'Team': 'IT (Portal-Admin)',
        'Lager-Nr': f'{len(sscc_aktiv)} Lager-Nrn (siehe Detail-Sheet)',
        'Diagnose': f'{len(sscc_aktiv)} Portal-Listings (status A oder M) sind veraltet: Geräte sind im AMM inzwischen Teil einer SSCC-Palette unter primärer Lager-Nr — Einzel-Listing im Portal stimmt nicht mehr',
        'Step 1': 'Detail-Sheet "SSCC-Mapping" öffnen — pro Eintrag steht die primäre SSCC-Palette + Status',
        'Step 2': 'Mit Sales klären: Palette als Bündel listen ODER Einzel-Listing entfernen',
        'Step 3': 'Portal-Update: Einzel-Lager-Nr aus Listing entfernen / SSCC-Bündel anlegen',
        'Tool': 'Portal-Admin · Detail-Sheet "SSCC-Mapping" · AMM-Admin',
        'Verifikation': 'Nächster Stock-Analysis-Export enthält die alten Einzel-Lager-Nrn NICHT mehr',
        'VK (€)': float(sscc_aktiv['Selling_Price'].sum()),
        'Alter (T)': int(sscc_aktiv['alter_tage'].median()) if sscc_aktiv['alter_tage'].notna().any() else 0,
        'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
    })

# IT-TASK C: Echte Geister Status-leer (28)
n_leer_echt = (echte_geister['status_norm']=='leer').sum()
it_tasks.append({
    'Prio': 'P2',
    'Typ': 'STATUS-KLAERUNG',
    'Team': 'IT (Portal-Admin)',
    'Lager-Nr': f'{n_leer_echt} echte Geister mit Status=leer',
    'Diagnose': f'{n_leer_echt} Portal-Listings mit Status=leer (NaN), nicht in AMM und auch nicht über LagerNr2 verknüpft — Bedeutung unklar',
    'Step 1': 'Mit Portal-Software-Hersteller klären: was bedeutet leerer Status-Wert?',
    'Step 2': 'Doku ins Wiki: Status-Konvention (A/M/leer)',
    'Step 3': f'Massen-Update: alle {n_leer_echt} nach Klärung entweder auf A setzen oder entfernen',
    'Tool': 'Portal-Admin · Backend-Doku',
    'Verifikation': 'Status-Verteilung im Stock-Export: nur noch A oder M, kein NaN',
    'VK (€)': float(echte_geister[echte_geister['status_norm']=='leer']['Selling_Price'].sum()),
    'Alter (T)': int(echte_geister[echte_geister['status_norm']=='leer']['alter_tage'].median()) if (echte_geister['status_norm']=='leer').any() else 0,
    'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
})

# Status-leer mit SSCC-Mapping (20) — eigene Task
n_leer_sscc = (sscc_mapping['status_norm']=='leer').sum()
if n_leer_sscc > 0:
    it_tasks.append({
        'Prio': 'P2',
        'Typ': 'SSCC-LISTING-VERALTET',
        'Team': 'IT (Portal-Admin)',
        'Lager-Nr': f'{n_leer_sscc} Status-leer Listings mit SSCC-Mapping',
        'Diagnose': f'{n_leer_sscc} Status-leer Portal-Listings sind im AMM als Teil einer SSCC-Palette verbucht — veraltete Einzel-Listings',
        'Step 1': 'Detail-Sheet "SSCC-Mapping" → Status-leer-Block',
        'Step 2': 'Einzel-Listings entfernen (Geräte sind palettiert)',
        'Step 3': 'Massen-Update via Portal-Admin-Backend',
        'Tool': 'Portal-Admin · Detail-Sheet',
        'Verifikation': 'Nächster Stock-Analysis-Export enthält diese Lager-Nrn nicht mehr',
        'VK (€)': float(sscc_mapping[sscc_mapping['status_norm']=='leer']['Selling_Price'].sum()),
        'Alter (T)': int(sscc_mapping[sscc_mapping['status_norm']=='leer']['alter_tage'].median()) if (sscc_mapping['status_norm']=='leer').any() else 0,
        'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
    })

# IT-TASK D: JTL-DOPPELBUCHUNG (Live-Beweis 900231005/006)
au_lookup_005 = jtl[jtl['Artikelnummer']=='900231005']['Bestell Nr.'].unique().tolist()
au_lookup_006 = jtl[jtl['Artikelnummer']=='900231006']['Bestell Nr.'].unique().tolist()
it_tasks.append({
    'Prio': 'P0',  # Allerhöchste — strukturelles Problem
    'Typ': 'DOPPELBUCHUNG-PREVENT',
    'Team': 'IT (JTL-Admin)',
    'Lager-Nr': '900231005, 900231006 (+ Risk-Universe)',
    'Diagnose': f'JTL erlaubt Mehrfach-Verwendung gleicher Lager-Nr: 900231005/006 in AU2026031399977 (13.03., versandt) UND AU2026060113692 (01.06., Johannes) — Live-Beweis aus Mail-Eskalation',
    'Step 1': 'JTL-Admin: prüfen ob Lager-Nr-Lock auf Auftrags-Position konfigurierbar (JTL-Wawi Einstellungen → Artikel → Eindeutigkeits-Constraints)',
    'Step 2': 'Wenn konfigurierbar: aktivieren + Test-Auftrag mit bereits gebuchter Lager-Nr',
    'Step 3': 'Wenn nicht: SQL-Trigger oder Custom-Plugin als Workaround. Alternativ: wöchentlicher Doppelbuchungs-Wächter (Skript)',
    'Tool': 'JTL-Wawi Konfiguration · SQL-Server · Custom-Plugin-Doku',
    'Verifikation': 'Test: zweiter Auftrag mit Lager-Nr aus offenem Auftrag → JTL muss Fehler werfen',
    'VK (€)': 0,
    'Alter (T)': 0,
    'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
})

# IT-TASK E: AMM↔JTL-Sync für AU2026060113692 (aus Mail)
it_tasks.append({
    'Prio': 'P1',
    'Typ': 'SYNC-LUECKE',
    'Team': 'IT (Schnittstelle JTL↔AMM)',
    'Lager-Nr': 'Auftrag AU2026060113692',
    'Diagnose': 'Auftrag in JTL angelegt (01.06.), nicht ans WMS (GBL) übertragen — Sebastian Reidl 01.06. 15:24: "liegt uns nicht vor"',
    'Step 1': 'Push-Mechanismus JTL→AMM-WMS-Schnittstelle prüfen: läuft der Job? Letzter erfolgreicher Push?',
    'Step 2': 'Auftrag AU2026060113692 manuell ans WMS pushen (Workaround)',
    'Step 3': 'Monitoring einrichten: Alarm wenn Auftrag >2h in JTL ohne WMS-Bestätigung',
    'Tool': 'JTL Worker-Service · AMM-Schnittstellen-Log · Monitoring-Tool',
    'Verifikation': 'Sebastian/GBL bestätigt: Auftrag sichtbar im WMS',
    'VK (€)': 0,
    'Alter (T)': 0,
    'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
})

# ============================================================
# LAGER-TASK-LISTE bauen
# ============================================================
lager_tasks = []

# Listing-Lücken (173) — Klassifizierung
qe_only_amm = b[(b['Status']=='QE') & (~b['Lager-Nr'].isin(s_set))].copy()
qe_only_amm['we_dt']=pd.to_datetime(qe_only_amm['WE'],errors='coerce',dayfirst=True)
sl = qe_only_amm['Lager-Nr'].isin(sold_set)
pal = qe_only_amm['Bez'].str.contains('Set Artikel|Unsortiert',case=False,na=False)
elhr = qe_only_amm['Lagerplatz'].astype(str).str.upper().str.startswith('ELHR')
fresh = qe_only_amm['we_dt'] > (NOW - pd.Timedelta(days=30))
echte_listing = qe_only_amm[~sl & ~pal & ~elhr & ~fresh].copy()
echte_listing['ber'] = echte_listing['Lagerplatz'].astype(str).str.extract(r'^([A-Z]+\d*)')[0]

# Gruppieren nach Lagerplatz-Bereich (effizientes Abarbeiten)
for bereich, grp in echte_listing.groupby('ber'):
    lager_tasks.append({
        'Prio': 'P1' if len(grp) >= 10 else 'P2',
        'Typ': 'LISTING-LUECKE',
        'Team': 'Lager (Klassifizierung)',
        'Lager-Nr': f'{len(grp)} Geräte in Bereich {bereich}',
        'Diagnose': f'{len(grp)} klassifizierte QE-Geräte im Hauptlager {bereich}, aber nicht im Portal gelistet — verlorener Umsatz',
        'Step 1': f'Rundgang Bereich {bereich}: alle Lager-Nrn aus Detail-Sheet physisch verifizieren (siehe Sheet "Detail Listing-Lücken")',
        'Step 2': 'Klassifizierungs-Status prüfen: Foto + Bewertung vollständig? Wenn nein: nacharbeiten',
        'Step 3': 'Ans Portal-Team übergeben: Excel mit Lager-Nr + Bewertung → Listing-Anlage',
        'Tool': 'Detail-Sheet (in dieser Excel) · Klassifizierungs-Tool · Foto-Kamera',
        'Verifikation': 'Nächster Stock-Analysis-Export enthält alle Lager-Nrn aus diesem Block',
        'VK (€)': 0,
        'Alter (T)': int((NOW - grp['we_dt'].min()).days) if grp['we_dt'].notna().any() else 0,
        'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
    })

# Set Artikel / Paletten (32) — Bündel-Listing-Konzept
n_pal = pal.sum()
lager_tasks.append({
    'Prio': 'P3',
    'Typ': 'PALETTEN-LISTING',
    'Team': 'Lager + Sales',
    'Lager-Nr': f'{n_pal} Sammelpaletten',
    'Diagnose': f'{n_pal} gemischte Sammelpaletten im QE-Status — aktuell nicht im Portal listbar weil Bündel',
    'Step 1': 'Mit Sales abstimmen: Konvolut-Konzept für Sammelposten — als Paket-Listing im Portal?',
    'Step 2': 'Inventur pro Palette: was steckt drin, geschätzter Wert?',
    'Step 3': 'Konvolute-Generator aus Cockpit nutzen (Reiter "Verkaufspriorität")',
    'Tool': 'Vertriebs-Cockpit · physische Palette',
    'Verifikation': 'Konvolut im Portal verkauft (Marker: in All-Sold mit Sammel-Notiz)',
    'VK (€)': 0,
    'Alter (T)': 0,
    'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
})

# Überzählig (105) — physische Inventur
# Überzählig: KILLCRITIC-korrigiert nach LagerNr2-Check (Mapping über SSCC entfernen)
n_uz_unknown = len(uz_lnrs - b_set - s_set - sold_set - jtl_lnrs - b_lnr2_set)
lager_tasks.append({
    'Prio': 'P2',
    'Typ': 'INVENTUR-UEBERZAEHLIG',
    'Team': 'Lager (Inventur)',
    'Lager-Nr': f'{n_uz_unknown} Lager-Nrn (nach SSCC-Bereinigung)',
    'Diagnose': f'{n_uz_unknown} Lager-Nrn aus Überzählig-Liste sind in KEINEM System (AMM, AMM-LagerNr2, Portal, Verkauf, JTL) — physisch wahrgenommen aber nicht erfasst',
    'Step 1': 'Physische Inventur: sind die Lager-Nrn wirklich vorhanden? Wo? (Siehe Sheet "Detail Überzählig")',
    'Step 2': 'Wenn JA: AMM-Anlage mit Status QE + Lagerplatz + Bewertung',
    'Step 3': 'Wenn NEIN: aus Überzählig-Liste entfernen (= geklärt)',
    'Tool': 'Detail-Sheet · AMM-Anlage-UI · physisches Lager',
    'Verifikation': 'Überzählig-Liste auf 0 reduziert + neue AMM-Einträge im nächsten BESTAND-Export',
    'VK (€)': 0,
    'Alter (T)': 0,
    'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
})

# ELHR-Umlager (84) — Umlager beschleunigen
elhr_count = elhr.sum()
lager_tasks.append({
    'Prio': 'P3',
    'Typ': 'UMLAGER-DURCHSATZ',
    'Team': 'Lager (Umlager)',
    'Lager-Nr': f'{elhr_count} Geräte in ELHR',
    'Diagnose': f'{elhr_count} Geräte hängen im Eingangs-/Umlagerbereich ELHR — verzögern Portal-Listing',
    'Step 1': 'Umlager-Stau analysieren: welche ELHR-Bereiche besonders voll? (siehe Detail-Sheet)',
    'Step 2': 'Bewegungs-Plan: Geräte aus ELHR raus an Verkaufsplatz',
    'Step 3': 'Wöchentliches Monitoring der ELHR-Geräte-Anzahl (sollte stabil <100 sein)',
    'Tool': 'Lager-Verwaltung · ELHR-Reporting',
    'Verifikation': 'ELHR-Bestand <100 Lager-Nrn, jeder Eintrag <14 Tage alt',
    'VK (€)': 0,
    'Alter (T)': 0,
    'Status': 'OFFEN', 'Bearbeiter': '', 'Datum erledigt': '', 'Notizen': '',
})

# === EXCEL ===
wb = Workbook()
INK='1d1d1f'; SUBTLE='595959'; HEAD_BG='1F4E79'
RED='C00000'; AMBER='D97706'; GREEN='2da14d'

# ============ SHEET 1: ÜBERSICHT ============
ws = wb.active
ws.title = 'Übersicht'
ws['A1'] = 'Team-Handlungsliste — IT + Lager'
ws['A1'].font = Font(size=18, bold=True, color=HEAD_BG)
ws.merge_cells('A1:G1')
ws['A2'] = f'Stand: {NOW.strftime("%d.%m.%Y")} · Quellen: BESTAND 01.06 · Stock 02.06 · All-Sold kombiniert · JTL 02.06 · Auftrag AU2026031399977 · Überzählig'
ws['A2'].font = Font(size=10, italic=True, color=SUBTLE)
ws.merge_cells('A2:G2')

# Team-Summary
ws['A4'] = 'TEAM-AUFTEILUNG'
ws['A4'].font = Font(size=12, bold=True, color=HEAD_BG)
ws['A4'].fill = PatternFill('solid', fgColor='DDEBF7')
ws.merge_cells('A4:G4')
ws.row_dimensions[4].height = 22

summary = [
    ('Team', 'P0', 'P1', 'P2', 'P3', 'Σ Tasks', 'Reiter'),
    ('🛠 IT (Portal-Admin, JTL-Admin, AMM-Admin, Schnittstelle)',
     sum(1 for t in it_tasks if t['Prio']=='P0'),
     sum(1 for t in it_tasks if t['Prio']=='P1'),
     sum(1 for t in it_tasks if t['Prio']=='P2'),
     sum(1 for t in it_tasks if t['Prio']=='P3'),
     len(it_tasks), '🛠 IT-Tasks'),
    ('📦 Lager (Klassifizierung, Inventur, Umlager)',
     sum(1 for t in lager_tasks if t['Prio']=='P0'),
     sum(1 for t in lager_tasks if t['Prio']=='P1'),
     sum(1 for t in lager_tasks if t['Prio']=='P2'),
     sum(1 for t in lager_tasks if t['Prio']=='P3'),
     len(lager_tasks), '📦 Lager-Tasks'),
]
for i, row in enumerate(summary, 5):
    for j, val in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=val)
        if i == 5:
            c.font = Font(size=10, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=HEAD_BG)
        else:
            c.font = Font(size=10)
            c.alignment = Alignment(vertical='center', wrap_text=(j==1))
            if j in (2,3,4,5,6):
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.font = Font(size=11, bold=True)
ws.row_dimensions[5].height = 22
ws.row_dimensions[6].height = 32
ws.row_dimensions[7].height = 32

# Priorisierungs-Legende
ws.cell(row=9, column=1, value='PRIORITÄTEN').font = Font(size=12, bold=True, color=HEAD_BG)
ws.cell(row=9, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
ws.merge_cells('A9:G9')
prio_legend = [
    ('P0 STRUKTURELL', 'Behebt System-Lücke, verhindert künftige Fälle (z.B. JTL-Lock)', 'C00000'),
    ('P1 AKUT', 'HEUTE/diese Woche: Storno-Risiko, kritische Sync-Lücke, Live-Verkaufsfall', 'D97706'),
    ('P2 WICHTIG', 'KW 23/24: Klärung, Listing-Lücken, Inventur-Korrektur', '7F6000'),
    ('P3 LAUFEND', 'Prozess-Verbesserung, kein direkter Schaden', '2DA14D'),
]
for i, (k, v, c) in enumerate(prio_legend, 10):
    ws.cell(row=i, column=1, value=k).font = Font(size=10, bold=True, color=c)
    ws.cell(row=i, column=2, value=v).font = Font(size=10)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)

# Roadmap
ws.cell(row=15, column=1, value='ROADMAP — Reihenfolge').font = Font(size=12, bold=True, color=HEAD_BG)
ws.cell(row=15, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
ws.merge_cells('A15:G15')
roadmap = [
    '1. HEUTE  · IT-Task DOPPELBUCHUNG-PREVENT (P0) → JTL-Admin: Lager-Nr-Lock konfigurieren ODER Workaround. Hebel: alle künftigen Doppelverkäufe verhindert.',
    '2. HEUTE  · IT-Task SYNC-LUECKE (P1) → AU2026060113692 manuell ans WMS pushen. Sebastian/GBL braucht den Auftrag.',
    '3. HEUTE  · IT-Tasks GEIST-AKTIV (14×) + AMM-LÜCKE (P1) → erst Lager-Bestätigung physisch, dann Portal-Sperre ODER AMM-Anlage.',
    '4. KW 23  · IT-Task GEIST-MANUELL (3×) + STATUS-KLAERUNG (47×). Portal-Admin-Klärung mit Software-Hersteller.',
    '5. KW 23-24 · Lager-Task LISTING-LUECKE (173 Geräte, gruppiert nach Bereich). Klassifizierung + Portal-Übergabe.',
    '6. KW 23-24 · Lager-Task INVENTUR-UEBERZAEHLIG (105 Lager-Nrn). Physisch verifizieren + AMM-Anlage oder Liste-Bereinigung.',
    '7. Laufend · Lager-Task UMLAGER-DURCHSATZ + PALETTEN-LISTING (Prozess-Themen).',
]
for i, e in enumerate(roadmap, 16):
    c = ws.cell(row=i, column=1, value=e)
    c.font = Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
    ws.row_dimensions[i].height = 28

# Spaltenbreiten
for col, w in zip('ABCDEFG', [60, 8, 8, 8, 8, 10, 18]):
    ws.column_dimensions[col].width = w

# ============ HELPER für Task-Sheets ============
TASK_HEADERS = ['Prio','Typ','Team','Lager-Nr','Diagnose','Step 1','Step 2','Step 3','Tool','Verifikation','VK (€)','Alter (T)','Status','Bearbeiter','Datum erledigt','Notizen']
TASK_WIDTHS = [6, 22, 24, 26, 50, 45, 45, 45, 32, 40, 11, 9, 13, 14, 14, 30]

def write_task_sheet(ws_t, title, tasks, color_header):
    ws_t['A1'] = title
    ws_t['A1'].font = Font(size=14, bold=True, color=color_header)
    ws_t.merge_cells('A1:P1')
    ws_t['A2'] = f'{len(tasks)} Tasks · Status zum Abhaken in Spalte M'
    ws_t['A2'].font = Font(size=10, italic=True, color=SUBTLE)
    ws_t.merge_cells('A2:P2')
    HR = 4
    for i, h in enumerate(TASK_HEADERS, 1):
        c = ws_t.cell(row=HR, column=i, value=h)
        c.font = Font(size=10.5, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=color_header)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_t.row_dimensions[HR].height = 32

    prio_color = {'P0':'C00000','P1':'D97706','P2':'7F6000','P3':'2DA14D'}
    for i, t in enumerate(tasks):
        row = HR + 1 + i
        for j, h in enumerate(TASK_HEADERS, 1):
            val = t.get(h, '')
            c = ws_t.cell(row=row, column=j, value=val)
            c.font = Font(size=9.5)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.border = Border(bottom=Side(style='thin', color='D9D9D9'))
            if h == 'Prio':
                c.font = Font(size=10, bold=True, color=prio_color.get(val, INK))
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif h == 'VK (€)':
                if val: c.number_format = '#,##0.00 €'; c.alignment = Alignment(horizontal='right', vertical='top')
            elif h in ('Alter (T)',):
                c.alignment = Alignment(horizontal='right', vertical='top')
            elif h == 'Status':
                c.font = Font(size=9.5, bold=True)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill('solid', fgColor='FFF2CC')
        ws_t.row_dimensions[row].height = 110

    # Status-Dropdown
    dv = DataValidation(type='list', formula1='"OFFEN,IN ARBEIT,WARTET,ERLEDIGT,IRRELEVANT"', allow_blank=True)
    dv.add(f'M{HR+1}:M{HR+len(tasks)}')
    ws_t.add_data_validation(dv)

    # Conditional Formatting Status
    ws_t.conditional_formatting.add(f'M{HR+1}:M{HR+len(tasks)}',
        FormulaRule(formula=[f'M{HR+1}="ERLEDIGT"'], stopIfTrue=False,
                    fill=PatternFill('solid', fgColor='C6EFCE'),
                    font=Font(color='006100', bold=True)))
    ws_t.conditional_formatting.add(f'M{HR+1}:M{HR+len(tasks)}',
        FormulaRule(formula=[f'M{HR+1}="IN ARBEIT"'], stopIfTrue=False,
                    fill=PatternFill('solid', fgColor='BDD7EE'),
                    font=Font(color='1F4E79', bold=True)))
    ws_t.conditional_formatting.add(f'M{HR+1}:M{HR+len(tasks)}',
        FormulaRule(formula=[f'M{HR+1}="OFFEN"'], stopIfTrue=False,
                    fill=PatternFill('solid', fgColor='FFE699'),
                    font=Font(color='7F6000', bold=True)))

    # Spaltenbreiten
    for j, w in enumerate(TASK_WIDTHS, 1):
        ws_t.column_dimensions[get_column_letter(j)].width = w
    ws_t.auto_filter.ref = f'A{HR}:{get_column_letter(len(TASK_HEADERS))}{HR+len(tasks)}'
    ws_t.freeze_panes = f'E{HR+1}'

# Sortieren nach Prio (P0 zuerst)
prio_order = {'P0':0,'P1':1,'P2':2,'P3':3}
it_tasks.sort(key=lambda t: (prio_order[t['Prio']], -t.get('VK (€)',0)))
lager_tasks.sort(key=lambda t: (prio_order[t['Prio']], -t.get('VK (€)',0)))

# Sheet IT
ws_it = wb.create_sheet('🛠 IT-Tasks')
write_task_sheet(ws_it, f'🛠 IT-TASKS — {len(it_tasks)} Aufgaben für Portal-Admin / JTL-Admin / AMM-Admin / Schnittstelle', it_tasks, RED)

# Sheet Lager
ws_lg = wb.create_sheet('📦 Lager-Tasks')
write_task_sheet(ws_lg, f'📦 LAGER-TASKS — {len(lager_tasks)} Aufgaben für Klassifizierung / Inventur / Umlager', lager_tasks, AMBER)

# ============ DETAIL: Listing-Lücken (für Lager) ============
ws_det1 = wb.create_sheet('Detail Listing-Lücken')
ws_det1['A1'] = f'Detail: 173 Listing-Lücken — Lager-Team Klassifizierung'
ws_det1['A1'].font = Font(size=14, bold=True, color=AMBER)
ws_det1.merge_cells('A1:H1')
ws_det1['A2'] = 'Gruppiert nach Bereich (Bereich-Manager kann seinen Block sofort abarbeiten)'
ws_det1['A2'].font = Font(size=10, italic=True, color=SUBTLE)
ws_det1.merge_cells('A2:H2')

dethdrs = ['Bereich','Lager-Nr','Bezeichnung','Lagerplatz','WE-Datum','Notiz','Auftrag-Bind','Status zum Abhaken']
for i, h in enumerate(dethdrs, 1):
    c = ws_det1.cell(row=4, column=i, value=h)
    c.font = Font(size=10.5, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=AMBER)
    c.alignment = Alignment(horizontal='left', vertical='center')
ws_det1.row_dimensions[4].height = 22

echte_sorted = echte_listing.sort_values(['ber','Bez']).reset_index(drop=True)
for i, r in echte_sorted.iterrows():
    row = 5 + i
    ws_det1.cell(row=row, column=1, value=r['ber'])
    ws_det1.cell(row=row, column=2, value=r['Lager-Nr']).font = Font(name='Consolas', size=10)
    ws_det1.cell(row=row, column=3, value=r['Bez'])
    ws_det1.cell(row=row, column=4, value=r['Lagerplatz'])
    ws_det1.cell(row=row, column=5, value=r['we_dt'].date() if pd.notna(r['we_dt']) else '')
    ws_det1.cell(row=row, column=5).number_format = 'DD.MM.YYYY'
    ws_det1.cell(row=row, column=6, value=r['Notiz'] if pd.notna(r['Notiz']) else '')
    ws_det1.cell(row=row, column=7, value=r['Auftrag'] if pd.notna(r['Auftrag']) else '')
    ws_det1.cell(row=row, column=8, value='OFFEN').fill = PatternFill('solid', fgColor='FFF2CC')

# Status-Dropdown
dv = DataValidation(type='list', formula1='"OFFEN,FOTO,KLASSIF,PORTAL,ERLEDIGT"', allow_blank=True)
dv.add(f'H5:H{4+len(echte_sorted)}')
ws_det1.add_data_validation(dv)
ws_det1.conditional_formatting.add(f'H5:H{4+len(echte_sorted)}',
    FormulaRule(formula=['H5="ERLEDIGT"'], stopIfTrue=False,
                fill=PatternFill('solid', fgColor='C6EFCE'), font=Font(color='006100', bold=True)))

for col, w in zip('ABCDEFGH', [10, 13, 32, 18, 12, 22, 18, 18]):
    ws_det1.column_dimensions[col].width = w
ws_det1.auto_filter.ref = f'A4:H{4+len(echte_sorted)}'
ws_det1.freeze_panes = 'A5'

# ============ DETAIL: Überzählig ============
ws_det2 = wb.create_sheet('Detail Überzählig')
ws_det2['A1'] = f'Detail: 105 Überzählig-Lager-Nrn — Lager-Team Inventur'
ws_det2['A1'].font = Font(size=14, bold=True, color=AMBER)
ws_det2.merge_cells('A1:D1')
ws_det2['A2'] = 'Lager-Nrn aus Überzählig-Liste, die in keinem System (AMM/Stock/AllSold/JTL) auffindbar sind'
ws_det2['A2'].font = Font(size=10, italic=True, color=SUBTLE)
ws_det2.merge_cells('A2:D2')

for i, h in enumerate(['Lager-Nr','Format','Physisch da?','Status zum Abhaken'], 1):
    c = ws_det2.cell(row=4, column=i, value=h)
    c.font = Font(size=10.5, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=AMBER)

import re
unknown_uz = sorted(uz_lnrs - b_set - s_set - sold_set - jtl_lnrs - b_lnr2_set)
for i, lnr in enumerate(unknown_uz):
    row = 5 + i
    ws_det2.cell(row=row, column=1, value=lnr).font = Font(name='Consolas', size=10)
    fmt = 'Standard 9-stellig' if re.match(r'^9\d{8}$',lnr) else 'Sonstige' if not re.match(r'^\d+$',lnr) else f'{len(lnr)}-stellig'
    ws_det2.cell(row=row, column=2, value=fmt)
    ws_det2.cell(row=row, column=3, value='').fill = PatternFill('solid', fgColor='FFF2CC')
    ws_det2.cell(row=row, column=4, value='OFFEN').fill = PatternFill('solid', fgColor='FFF2CC')

dv2 = DataValidation(type='list', formula1='"OFFEN,GESUCHT,GEFUNDEN,NICHT-DA,IM-AMM-ANGELEGT,ERLEDIGT"', allow_blank=True)
dv2.add(f'D5:D{4+len(unknown_uz)}')
ws_det2.add_data_validation(dv2)

for col, w in zip('ABCD', [16, 22, 14, 20]):
    ws_det2.column_dimensions[col].width = w
ws_det2.auto_filter.ref = f'A4:D{4+len(unknown_uz)}'
ws_det2.freeze_panes = 'A5'

# ============ DETAIL: SSCC-Mapping (29) ============
ws_sscc = wb.create_sheet('Detail SSCC-Mapping')
ws_sscc['A1'] = f'Detail: {len(sscc_mapping)} SSCC-Mapping-Fälle — Portal listet veraltete Einzel-Lager-Nr'
ws_sscc['A1'].font = Font(size=14, bold=True, color=RED)
ws_sscc.merge_cells('A1:J1')
ws_sscc['A2'] = 'Gerät ist im AMM als Teil einer SSCC-Palette geführt (primäre 18-stellige Lager-Nr). Portal listet noch die alte 9-stellige Einzel-Lager-Nr.'
ws_sscc['A2'].font = Font(size=10, italic=True, color=SUBTLE)
ws_sscc.merge_cells('A2:J2')

sscc_headers = ['Portal-Lager-Nr','Portal-Status','Marke','Modell','VK (€)','Alter (T)','AMM-Primär-Lager-Nr (SSCC)','AMM-Status','AMM-Bez','Status zum Abhaken']
for i, h in enumerate(sscc_headers, 1):
    c = ws_sscc.cell(row=4, column=i, value=h)
    c.font = Font(size=10.5, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=RED)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws_sscc.row_dimensions[4].height = 28

sscc_sorted = sscc_mapping.sort_values(['status_norm','Selling_Price'], ascending=[True, False]).reset_index(drop=True)
for i, r in sscc_sorted.iterrows():
    row = 5 + i
    ln = r['lager_number']
    amm_info = b_sscc_lookup.get(ln, ('?', '?', '?', '?'))
    ws_sscc.cell(row=row, column=1, value=ln).font = Font(name='Consolas', size=10)
    ws_sscc.cell(row=row, column=2, value=r['status_norm']).font = Font(size=10, bold=True)
    ws_sscc.cell(row=row, column=3, value=r['brand'] if pd.notna(r['brand']) else '')
    ws_sscc.cell(row=row, column=4, value=str(r['model'])[:40] if pd.notna(r['model']) else '')
    c5 = ws_sscc.cell(row=row, column=5, value=float(r['Selling_Price']))
    c5.number_format = '#,##0.00 €'; c5.alignment = Alignment(horizontal='right')
    ws_sscc.cell(row=row, column=6, value=int(r['alter_tage']) if pd.notna(r['alter_tage']) else 0).alignment = Alignment(horizontal='right')
    ws_sscc.cell(row=row, column=7, value=amm_info[0]).font = Font(name='Consolas', size=9)
    ws_sscc.cell(row=row, column=8, value=amm_info[1])
    ws_sscc.cell(row=row, column=9, value=amm_info[2])
    ws_sscc.cell(row=row, column=10, value='OFFEN').fill = PatternFill('solid', fgColor='FFF2CC')

dv_sscc = DataValidation(type='list', formula1='"OFFEN,GEKLAERT-PALETTE-LISTEN,LISTING-ENTFERNT,ERLEDIGT"', allow_blank=True)
dv_sscc.add(f'J5:J{4+len(sscc_sorted)}')
ws_sscc.add_data_validation(dv_sscc)
ws_sscc.conditional_formatting.add(f'J5:J{4+len(sscc_sorted)}',
    FormulaRule(formula=['J5="ERLEDIGT"'], stopIfTrue=False,
                fill=PatternFill('solid', fgColor='C6EFCE'), font=Font(color='006100', bold=True)))

for col, w in zip('ABCDEFGHIJ', [14, 11, 12, 32, 11, 9, 22, 11, 26, 22]):
    ws_sscc.column_dimensions[col].width = w
ws_sscc.auto_filter.ref = f'A4:J{4+len(sscc_sorted)}'
ws_sscc.freeze_panes = 'A5'

# ============ METHODIK ============
ws_m = wb.create_sheet('Methodik')
ws_m['A1'] = 'Methodik & Quellen'
ws_m['A1'].font = Font(size=14, bold=True, color=HEAD_BG)

content = [
    ('', ''),
    ('Quellen (Snapshot 01./02.06.2026)', ''),
    ('BESTAND (AMM)', BEST),
    ('Stock-Analysis (Portal)', STOCK),
    ('All-Sold (Portal Verkaufshistorie)', 'kombiniert Jan-Jun 2026'),
    ('JTL-Auftrags-Export', f'{USER}/elvinci - JTL EXPORT/JTL-Export-Aufträge-02062026.csv'),
    ('Auftrag AU2026031399977 (CSV)', ORDER_CSV),
    ('Überzählig-Liste', UZ),
    ('', ''),
    ('Status-Konvention je Task', ''),
    ('OFFEN', 'Noch nicht begonnen'),
    ('IN ARBEIT', 'Bearbeiter weist sich zu, Datum laufend'),
    ('WARTET', 'Externe Klärung nötig (z.B. Portal-Hersteller)'),
    ('ERLEDIGT', 'Verifikation bestätigt, Task abgeschlossen'),
    ('IRRELEVANT', 'Nach Klärung obsolet (z.B. AMM-Korrektur war schon erfolgt)'),
    ('', ''),
    ('Verifikation: jeder Task hat eine PRÜFBARE Erfolgs-Bedingung', ''),
    ('Nicht abhaken ohne Verifikation', 'Sonst entsteht Schein-Fortschritt'),
    ('', ''),
    ('KILLCRITIC-Korrektur Version 2 (SSCC-Mapping)', ''),
    ('Erkenntnis', 'BESTAND hat zwei Lager-Nr-Spalten: primär (Lager-Nr) + sekundär (LagerNr2)'),
    ('Frühere Schwäche', 'Diff-Analysen prüften nur primär → 29 von 65 Geistern waren in Wahrheit SSCC-Mapping'),
    ('Korrektur', 'Geister-Klassifikation jetzt Multi-Spalten: echte Geister nur wenn auch nicht in LagerNr2'),
    ('Folge', 'Statt 14 echte AKTIV-Geister: nur 5. Statt 48 Status-leer: nur 28. NEUE Kategorie SSCC-LISTING-VERALTET (29).'),
]
for i, (k, v) in enumerate(content, 3):
    ws_m.cell(row=i, column=1, value=k).font = Font(size=10, bold=k.startswith(('Quellen','Status','Verifikation')))
    ws_m.cell(row=i, column=2, value=v).font = Font(size=10)
ws_m.column_dimensions['A'].width = 40
ws_m.column_dimensions['B'].width = 90

wb.save(OUT)
print(f'\nExcel: {OUT}')
print(f'  IT-Tasks: {len(it_tasks)} (P0={sum(1 for t in it_tasks if t["Prio"]=="P0")}, P1={sum(1 for t in it_tasks if t["Prio"]=="P1")}, P2={sum(1 for t in it_tasks if t["Prio"]=="P2")})')
print(f'  Lager-Tasks: {len(lager_tasks)} (P1={sum(1 for t in lager_tasks if t["Prio"]=="P1")}, P2={sum(1 for t in lager_tasks if t["Prio"]=="P2")}, P3={sum(1 for t in lager_tasks if t["Prio"]=="P3")})')
print(f'  Detail Listing-Lücken: 173 Einzelgeräte')
print(f'  Detail Überzählig: {len(unknown_uz)} Lager-Nrn')
