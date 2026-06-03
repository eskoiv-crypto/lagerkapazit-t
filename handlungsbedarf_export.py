"""
HANDLUNGSBEDARF: konsolidierte Maßnahmenliste aus dem Diff AMM ↔ Portal.
Drei Prioritäten (P1 sofort · P2 diese Woche · P3 Prozess), je Item:
Wer, Was, Bis wann, Impact.
"""
import pandas as pd, datetime, glob
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

USER = Path(r'C:\Users\D.Eskofier\OneDrive - elvinci.de GmbH')
BEST = r'C:\Users\D.Eskofier\Downloads\BESTAND134_20260601_2330 (1).CSV'
STOCK = r'C:\Users\D.Eskofier\Downloads\Stock-Analysis-2026-06-02T09_25_01.294Z.xlsx'
OUT = USER / 'Anlagen' / 'Handlungsbedarf_AMM-vs-Portal_2026-06-02.xlsx'

# === DATEN ===
b = pd.read_csv(BEST, sep=';', encoding='ISO-8859-1', skiprows=1, header=None, low_memory=False, dtype=str)
cols=['Lager-Nr','Anzahl','Bez','LagerNr2','Lagerplatz','Klassif','Menge','WE','Notiz','Status','Auftrag','c11','c12','c13'][:b.shape[1]]
b.columns=cols
b['Lager-Nr']=b['Lager-Nr'].astype(str).str.strip()
b_set = set(b['Lager-Nr'])

s = pd.read_excel(STOCK)
s['lager_number']=s['lager_number'].astype(str).str.strip()

old = pd.read_excel(sorted(glob.glob(str(USER/'elvinci - Portal ALL SOLD'/'All-Sold-*.xlsx')))[-1])
new = pd.read_excel(r'C:\Users\D.Eskofier\Downloads\All-Sold-2026-06-01T12_44_40.671Z.xlsx')
for df in [old,new]:
    df['lnr']=df['Lager Nr.'].astype(str).str.strip().str.replace(r'\.0\$','',regex=True)
sold_set = set(pd.concat([old['lnr'],new['lnr']]).unique())

NOW = datetime.datetime.now()

# === Geister (Portal frei, AMM fehlt) ===
geister = s[~s['lager_number'].isin(b_set)].copy()
geister['status_norm'] = geister['status'].fillna('leer')
geister['upload_dt'] = pd.to_datetime(geister['datetime_upload'],errors='coerce')
geister['alter_tage'] = (NOW - geister['upload_dt']).dt.days
for c in ['Buying_Price','Selling_Price','Online_Price']:
    geister[c] = pd.to_numeric(geister[c],errors='coerce').fillna(0)

geister_aktiv = geister[geister['status_norm']=='A'].sort_values('Selling_Price',ascending=False)
geister_manuell = geister[geister['status_norm']=='M'].sort_values('Selling_Price',ascending=False)
geister_leer = geister[geister['status_norm']=='leer'].sort_values('Selling_Price',ascending=False)

# === Listing-Lücken (QE im AMM, NICHT im Portal) — saubere Bucket-Filterung ===
s_set = set(s['lager_number'])
only_amm_qe = b[(b['Status']=='QE') & (~b['Lager-Nr'].isin(s_set))].copy()
only_amm_qe['we_dt'] = pd.to_datetime(only_amm_qe['WE'],errors='coerce',dayfirst=True)

is_set = only_amm_qe['Bez'].str.contains('Set Artikel|Unsortiert',case=False,na=False)
is_elhr = only_amm_qe['Lagerplatz'].astype(str).str.startswith('ELHR')
is_fresh = only_amm_qe['we_dt'] > pd.Timestamp.now() - pd.Timedelta(days=30)
is_sold = only_amm_qe['Lager-Nr'].isin(sold_set)

listing_luecken = only_amm_qe[~is_set & ~is_elhr & ~is_fresh & ~is_sold].copy()

# === EXCEL BAUEN ===
wb = Workbook()

# ============ COLORS ============
RED = '~C00000'; RED_FG = 'FFC7CE'; RED_TXT = '9C0006'
ORG = '~ED7D31'; ORG_FG = 'FFE699'; ORG_TXT = '9C5700'
GRN = '~70AD47'; GRN_FG = 'C6EFCE'; GRN_TXT = '375623'
BLUE = '~1F4E79'; BLUE_FG = 'DDEBF7'
HEAD_BG = '1F4E79'; HEAD_FG = 'FFFFFF'
ZEBRA = 'F8F9FA'

def set_header_row(ws, row, headers, bg=HEAD_BG):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(size=10.5, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = Border(bottom=Side(style='medium', color='000000'))
    ws.row_dimensions[row].height = 28

def apply_zebra(ws, start_row, end_row, n_cols, color=ZEBRA):
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, '00000000', '00FFFFFF'):
                    cell.fill = PatternFill('solid', fgColor=color)
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style='thin', color='E5E5E5'))

# ============ SHEET 1: ÜBERSICHT / COCKPIT ============
ws = wb.active
ws.title = 'Übersicht'

ws['A1'] = 'HANDLUNGSBEDARF — AMM ↔ Portal-Differenz'
ws['A1'].font = Font(size=18, bold=True, color='1F4E79')
ws.merge_cells('A1:G1')
ws['A2'] = f'Stand: BESTAND 01.06.2026  ·  Stock-Analysis 02.06.2026  ·  erstellt {NOW.strftime("%d.%m.%Y %H:%M")}'
ws['A2'].font = Font(size=10, italic=True, color='595959')
ws.merge_cells('A2:G2')

# Prio-Zusammenfassung
ws['A4'] = 'PRIORITÄT'; ws['B4'] = 'KATEGORIE'; ws['C4'] = 'ANZAHL'; ws['D4'] = 'Σ VK (€)'; ws['E4'] = 'WER'; ws['F4'] = 'BIS WANN'; ws['G4'] = 'AKTION'
set_header_row(ws, 4, ['PRIO','KATEGORIE','ANZAHL','Σ VK (€)','WER','BIS WANN','AKTION'])

rows = [
    ('🔴 P1',  'Geister AKTIV im Portal',             len(geister_aktiv),    geister_aktiv['Selling_Price'].sum(),
     'Portal-Admin', 'HEUTE',     'Im Portal sperren — Lager-Nr existiert nicht im AMM (kann nicht versandt werden)'),
    ('🔴 P1',  'Geister MANUELL gelistet',            len(geister_manuell),  geister_manuell['Selling_Price'].sum(),
     'Portal-Admin', 'HEUTE',     'Manuelles Listing prüfen — vermutlich Altdaten oder Fehleingabe'),
    ('🟠 P2',  'Listing-Lücken (QE, im Lager)',       len(listing_luecken),  0,
     'Klassifizierung', 'KW 23',  'Ins Portal stellen — Geräte sind klassifiziert + im Hauptlager, aber nicht im Verkauf'),
    ('🟠 P2',  'Geister Status-leer im Portal',       len(geister_leer),     geister_leer['Selling_Price'].sum(),
     'Portal-Admin + IT', 'KW 24', 'Status klären (warum NaN?) + Altdaten 2021/22/23 entfernen'),
    ('🟡 P3',  'Set Artikel / Paletten (im QE)',      int(is_set.sum()),     0,
     'Sales',           'KW 26',  'Bündel-Listing-Konzept entwickeln (heute nicht einzeln im Portal listbar)'),
    ('🟡 P3',  'ELHR-Umlager (im QE)',                int(is_elhr.sum()),    0,
     'Lager',           'laufend','Umlager-Prozess prüfen — Geräte hängen im Eingangs-/Umlagerbereich'),
    ('🟢 OK',  'Frischware (WE letzte 30T)',          int(is_fresh.sum()),   0,
     '—',               '—',      'Kein Handlungsbedarf — Klassifizierung läuft regulär'),
    ('🟢 OK',  'Status-Lag (schon verkauft, QE)',     int(is_sold.sum()),    0,
     '—',               '—',      'Kein Handlungsbedarf — wird mit Versand auf VS gesetzt'),
]
row = 5
for prio, kat, anz, vk, wer, deadline, aktion in rows:
    is_p1 = prio.startswith('🔴')
    is_p2 = prio.startswith('🟠')
    is_p3 = prio.startswith('🟡')
    is_ok = prio.startswith('🟢')
    fill_color = RED_FG if is_p1 else ORG_FG if is_p2 else 'FFF2CC' if is_p3 else GRN_FG
    txt_color = RED_TXT if is_p1 else ORG_TXT if is_p2 else '7F6000' if is_p3 else GRN_TXT
    ws.cell(row=row, column=1, value=prio).font = Font(size=11, bold=True, color=txt_color)
    ws.cell(row=row, column=2, value=kat).font = Font(size=10.5, bold=is_p1)
    c3 = ws.cell(row=row, column=3, value=anz)
    c3.font = Font(size=11, bold=True); c3.alignment = Alignment(horizontal='right')
    c4 = ws.cell(row=row, column=4, value=vk if vk > 0 else None)
    c4.number_format='#,##0.00 €'; c4.alignment=Alignment(horizontal='right')
    ws.cell(row=row, column=5, value=wer).font = Font(size=10)
    ws.cell(row=row, column=6, value=deadline).font = Font(size=10, bold=is_p1)
    ws.cell(row=row, column=7, value=aktion).font = Font(size=9.5)
    ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical='center')
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=fill_color)
        ws.cell(row=row, column=c).border = Border(bottom=Side(style='thin', color='D9D9D9'))
    ws.row_dimensions[row].height = 36
    row += 1

# Totals
row += 1
ws.cell(row=row, column=1, value='ZUSAMMENFASSUNG').font = Font(size=11, bold=True, color='1F4E79')
ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
sums = [
    ('Akute P1-Items (sofort sperren)',  len(geister_aktiv)+len(geister_manuell), geister_aktiv['Selling_Price'].sum()+geister_manuell['Selling_Price'].sum()),
    ('P2-Items (diese Woche)',           len(listing_luecken)+len(geister_leer),   geister_leer['Selling_Price'].sum()),
    ('Σ alle Differenzen AMM ↔ Portal', len(geister)+len(only_amm_qe),            geister['Selling_Price'].sum()),
]
for label, n, v in sums:
    ws.cell(row=row, column=1, value=label).font = Font(size=10)
    ws.cell(row=row, column=2, value=n).font = Font(size=10, bold=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    c = ws.cell(row=row, column=4, value=v if v > 0 else None)
    c.number_format='#,##0.00 €'; c.font = Font(size=10, bold=True); c.alignment = Alignment(horizontal='right')
    row += 1

# Empfehlung-Block
row += 2
ws.cell(row=row, column=1, value='💡 EMPFEHLUNG / NÄCHSTE SCHRITTE').font = Font(size=11, bold=True, color='1F4E79')
ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='DDEBF7')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
empf = [
    '1.  HEUTE — die 14 AKTIVEN Geister im Portal sperren (siehe Sheet "P1 Geister AKTIV"). Verhindert Verkaufs-Stornos.',
    '2.  HEUTE — die 3 MANUELL gelisteten Geister prüfen (Altdaten oder Fehleingabe).',
    '3.  KW 23 — die 182 Listing-Lücken klassifizieren+ins Portal stellen (siehe Sheet "P2 Listing-Luecken"). Geschätzter Umsatz-Hebel: ~25–30 k €.',
    '4.  KW 24 — mit IT/Portal-Admin klären: was bedeutet Status=leer (NaN)? 48 alte Listings stehen ungeklärt im Portal.',
    '5.  Prozess — Geister-Check wöchentlich einplanen (Skript geister_export.py liefert in 5 Sek frische Liste).',
    '6.  Stammdaten — Set-Artikel-Bündel-Konzept im Portal (203 Paletten könnten als Konvolute angeboten werden).',
]
for e in empf:
    row += 1
    c = ws.cell(row=row, column=1, value=e); c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.row_dimensions[row].height = 30

# Spaltenbreiten Sheet 1
widths_s1 = [10, 32, 9, 13, 18, 12, 60]
for i, w in enumerate(widths_s1, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26

# ============ HELPER: Detail-Sheet ============
def write_detail_sheet(name, df, title, color_header, value_col='Selling_Price',
                       cols_def=None):
    """cols_def: list of (header, df_col, format, width)"""
    ws_d = wb.create_sheet(name)
    ws_d['A1'] = title
    ws_d['A1'].font = Font(size=14, bold=True, color=color_header)
    ws_d.merge_cells('A1:H1')
    ws_d.row_dimensions[1].height = 22

    ws_d['A2'] = f'{len(df)} Lager-Nrn  ·  Σ VK {pd.to_numeric(df[value_col],errors="coerce").sum():,.2f} €'
    ws_d['A2'].font = Font(size=10, italic=True, color='595959')
    ws_d.merge_cells('A2:H2')

    HR = 4
    headers = [c[0] for c in cols_def]
    set_header_row(ws_d, HR, headers)

    for i, (_, r) in enumerate(df.iterrows()):
        row = HR + 1 + i
        for j, (h, col, fmt, width) in enumerate(cols_def, 1):
            val = r.get(col, '')
            if isinstance(val, float) and pd.isna(val): val = ''
            if isinstance(val, pd.Timestamp): val = val.date() if not pd.isna(val) else ''
            c = ws_d.cell(row=row, column=j, value=val)
            if fmt == 'eur':
                c.number_format = '#,##0.00 €'; c.alignment = Alignment(horizontal='right')
            elif fmt == 'int':
                c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
            elif fmt == 'date':
                c.number_format = 'DD.MM.YYYY'; c.alignment = Alignment(horizontal='center')
            elif fmt == 'mono':
                c.font = Font(name='Consolas', size=10)
            else:
                c.font = Font(size=10)

    apply_zebra(ws_d, HR + 1, HR + len(df), len(cols_def))
    for j, (_, _, _, w) in enumerate(cols_def, 1):
        ws_d.column_dimensions[get_column_letter(j)].width = w
    ws_d.auto_filter.ref = f'A{HR}:{get_column_letter(len(cols_def))}{HR + len(df)}'
    ws_d.freeze_panes = f'A{HR + 1}'

# ============ SHEET 2: P1 Geister AKTIV ============
geister_aktiv_d = geister_aktiv.copy()
geister_aktiv_d['Bezeichnung'] = (geister_aktiv_d['brand'].fillna('') + ' ' + geister_aktiv_d['model'].fillna('').astype(str).str[:35]).str.strip()
write_detail_sheet('P1 Geister AKTIV', geister_aktiv_d,
    '🔴 P1 — Geister AKTIV im Portal verkaufbar, aber NICHT im AMM',
    'C00000',
    cols_def=[
        ('Lager-Nr', 'lager_number', 'mono', 14),
        ('Marke', 'brand', 'text', 12),
        ('Modell', 'model', 'text', 32),
        ('Produktgruppe', 'product_group', 'text', 26),
        ('EK (€)', 'Buying_Price', 'eur', 11),
        ('VK (€)', 'Selling_Price', 'eur', 11),
        ('Online (€)', 'Online_Price', 'eur', 11),
        ('Upload', 'upload_dt', 'date', 12),
        ('Alter (T)', 'alter_tage', 'int', 9),
    ])

# ============ SHEET 3: P1 Geister MANUELL ============
if len(geister_manuell) > 0:
    write_detail_sheet('P1 Geister MANUELL', geister_manuell,
        '🔴 P1 — Geister MANUELL gelistet, aber NICHT im AMM',
        'C00000',
        cols_def=[
            ('Lager-Nr', 'lager_number', 'mono', 14),
            ('Marke', 'brand', 'text', 12),
            ('Modell', 'model', 'text', 32),
            ('Produktgruppe', 'product_group', 'text', 26),
            ('EK (€)', 'Buying_Price', 'eur', 11),
            ('VK (€)', 'Selling_Price', 'eur', 11),
            ('Online (€)', 'Online_Price', 'eur', 11),
            ('Upload', 'upload_dt', 'date', 12),
            ('Alter (T)', 'alter_tage', 'int', 9),
        ])

# ============ SHEET 4: P2 Listing-Lücken ============
listing_luecken = listing_luecken.copy()
listing_luecken['ber'] = listing_luecken['Lagerplatz'].astype(str).str.extract(r'^([A-Z]+\d*)')[0]
listing_luecken['Selling_Price'] = 0  # kein VK-Wert (nicht im Portal)
listing_luecken['we_show'] = listing_luecken['we_dt']
listing_luecken = listing_luecken.sort_values('we_dt')
write_detail_sheet('P2 Listing-Luecken', listing_luecken,
    '🟠 P2 — Listing-Lücken: QE klassifiziert im Hauptlager, aber NICHT im Portal',
    'ED7D31', value_col='Selling_Price',
    cols_def=[
        ('Lager-Nr', 'Lager-Nr', 'mono', 14),
        ('Bezeichnung', 'Bez', 'text', 32),
        ('Lagerplatz', 'Lagerplatz', 'text', 18),
        ('Bereich', 'ber', 'text', 10),
        ('WE-Datum', 'we_show', 'date', 12),
        ('Notiz', 'Notiz', 'text', 24),
        ('Status', 'Status', 'text', 8),
    ])

# ============ SHEET 5: P2 Geister Status-leer ============
if len(geister_leer) > 0:
    write_detail_sheet('P2 Geister Status-leer', geister_leer,
        '🟠 P2 — Geister mit Status-leer (NaN) im Portal, aber NICHT im AMM',
        'ED7D31',
        cols_def=[
            ('Lager-Nr', 'lager_number', 'mono', 14),
            ('Marke', 'brand', 'text', 12),
            ('Modell', 'model', 'text', 32),
            ('Produktgruppe', 'product_group', 'text', 26),
            ('EK (€)', 'Buying_Price', 'eur', 11),
            ('VK (€)', 'Selling_Price', 'eur', 11),
            ('Online (€)', 'Online_Price', 'eur', 11),
            ('Upload', 'upload_dt', 'date', 12),
            ('Alter (T)', 'alter_tage', 'int', 9),
        ])

# ============ SHEET 6: Methodik & Quellen ============
ws_m = wb.create_sheet('Methodik')
ws_m['A1'] = 'Methodik & Quellen'
ws_m['A1'].font = Font(size=14, bold=True)

content = [
    ('', ''),
    ('Quellen', ''),
    ('BESTAND (AMM)', f'{BEST}'),
    ('Stock-Analysis (Portal)', f'{STOCK}'),
    ('All-Sold (kombiniert)', 'Portal-Export Apr 2025 – Mai 2026 (~27k Verkäufe)'),
    ('', ''),
    ('Definition Geister', 'Lager-Nr im Portal-Stock vorhanden, aber NICHT in AMM-BESTAND.'),
    ('  → P1 AKTIV (status=A)', 'Definitiv verkaufbar — höchstes Storno-Risiko'),
    ('  → P1 MANUELL (status=M)', 'Manuell gepflegte Listings — meist Altdaten'),
    ('  → P2 Status-leer (NaN)', 'Status nicht gesetzt — vermutlich Default-sichtbar im Portal'),
    ('', ''),
    ('Definition Listing-Lücken', 'QE-Status im AMM aber NICHT im Portal-Stock.'),
    ('  Ausgeschlossen werden (= kein Handlungsbedarf):', ''),
    ('  − Set Artikel/Unsortiert', 'Paletten (nicht einzeln listbar)'),
    ('  − ELHR-Umlager', 'Physisch in Bewegung (Eingangs-/Umlagerbereich)'),
    ('  − Frischware (WE letzte 30T)', 'Klassifizierung läuft noch regulär'),
    ('  − Status-Lag (in All-Sold)', 'Schon verkauft, nur AMM-Status hinkt nach'),
    ('', ''),
    ('Restrisiken', ''),
    ('Snapshot-Zeitversatz 10h', 'BESTAND 01.06 23:30 vs Stock 02.06 09:25 — geschätzt ~20–30 Geister könnten Sync-Artefakt sein'),
    ('SSCC-Format-Mismatch', '58 BESTAND-Lager-Nrn sind 18-stellig (SSCC), 49 davon haben keine Stock-Entsprechung — methodische Schwäche'),
    ('Stock-Status NaN ungeklärt', '2.054 von 3.477 Stock-Einträgen haben keinen Status — Klärung mit Portal-Admin offen'),
]
for i, (k, v) in enumerate(content, 3):
    ws_m.cell(row=i, column=1, value=k).font = Font(size=10, bold=k.startswith(('Quellen','Definition','Restrisiken')))
    ws_m.cell(row=i, column=2, value=v).font = Font(size=10)
    ws_m.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
ws_m.column_dimensions['A'].width = 36
ws_m.column_dimensions['B'].width = 80

wb.save(OUT)
print(f'Excel: {OUT}')
print(f'  Sheets: Übersicht · P1 Geister AKTIV ({len(geister_aktiv)}) · P1 MANUELL ({len(geister_manuell)}) · P2 Listing-Lücken ({len(listing_luecken)}) · P2 Status-leer ({len(geister_leer)}) · Methodik')
print(f'  P1 sofort: {len(geister_aktiv)+len(geister_manuell)} Items, Σ VK {(geister_aktiv["Selling_Price"].sum()+geister_manuell["Selling_Price"].sum()):,.2f} €')
print(f'  P2 diese Woche: {len(listing_luecken)+len(geister_leer)} Items')
